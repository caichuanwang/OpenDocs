from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal
from zipfile import BadZipFile, ZipFile, ZipInfo

from defusedxml import ElementTree as DefusedET
from defusedxml.common import DefusedXmlException
from openpyxl.utils.cell import get_column_letter, range_boundaries
from PIL import Image, ImageDraw, ImageFont

from opendocs._models import (
    Block,
    HeadingBlock,
    InlineText,
    ParagraphBlock,
    TableBlock,
    WarningRecord,
)
from opendocs.errors import CorruptDocumentError, LimitExceededError
from opendocs.parsers.xlsx.models import XlsxChartSlot, XlsxDocument, XlsxImageSlot
from opendocs.parsers.xlsx.preflight import (
    MAX_CHART_CACHE_POINTS,
    MAX_DRAWING_OBJECTS,
    XlsxPreflight,
    XlsxPreflightSheet,
    _read_relationships,
    _relationships_part,
)
from opendocs.vision.base import VisionRequest, VisionRequestKind
from opendocs.vision.images import PreparedImage, prepare_image

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWING_RELATIONSHIP = f"{_OFFICE_REL_NS}/drawing"
_CHART_RELATIONSHIP = f"{_OFFICE_REL_NS}/chart"
_IMAGE_RELATIONSHIP = f"{_OFFICE_REL_NS}/image"
_RELATIONSHIP_ID = f"{{{_OFFICE_REL_NS}}}id"
_RELATIONSHIP_EMBED = f"{{{_OFFICE_REL_NS}}}embed"

_CHART_TYPES = {
    "lineChart": "line",
    "barChart": "bar",
    "pieChart": "pie",
    "doughnutChart": "doughnut",
    "scatterChart": "scatter",
}
_AXIS_TYPES = {"catAx", "valAx", "dateAx", "serAx"}
_LOCAL_RANGE = re.compile(
    r"^(?P<sheet>'(?:[^']|'')+'|[^'!\[\]]+)!"
    r"(?P<start>\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6})"
    r"(?::(?P<end>\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}))?$",
    re.IGNORECASE,
)
_SAFE_ARTIFACT_SUFFIX = re.compile(r"^\.[A-Za-z0-9]{1,8}$")
_PREVIEW_WIDTH = 1_280
_PREVIEW_PADDING = 40
_PREVIEW_LINE_HEIGHT = 26
_PREVIEW_MAX_LINES = 80
_PREVIEW_MAX_LINE_CHARS = 180
_PREVIEW_SERIES_POINTS = 24

XLSX_CHART_VISION_PROMPT = (
    "这是由 XLSX 原生图表事实生成的语义卡片, 不是 Excel 外观还原。"
    "仅补充可由卡片支持的趋势、关系、标注和含义; 不要改写原生数值, "
    "不要猜测缺失数据, 并将结论明确标记为 '视觉解释'。"
)
XLSX_IMAGE_VISION_PROMPT = (
    "仅解释图片中可见的文字、标注、关系和含义; 涉及趋势时只描述可见证据, "
    "不要猜测未显示的内容, 并将结论明确标记为 '视觉解释'。"
)


@dataclass(frozen=True, slots=True)
class XlsxChartSeriesFacts:
    name: str
    categories: tuple[str, ...]
    values: tuple[str, ...]
    x_values: tuple[str, ...]
    y_values: tuple[str, ...]
    formulas: tuple[str, ...]
    unresolved_formulas: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class XlsxChartFacts:
    chart_type: str
    title: str
    axis_titles: tuple[str, ...]
    axis_labels: tuple[str, ...]
    data_labels: tuple[str, ...]
    formulas: tuple[str, ...]
    unresolved_formulas: tuple[str, ...]
    series: tuple[XlsxChartSeriesFacts, ...]


@dataclass(frozen=True, slots=True)
class XlsxVisualOccurrence:
    kind: Literal["chart", "image"]
    anchor: str
    blocks: tuple[Block, ...]
    artifact_name: str | None
    content_sha256: str | None
    alt_text: str | None
    object_name: str | None
    title: str | None


@dataclass(frozen=True, slots=True)
class XlsxVisualObjects:
    by_sheet: tuple[tuple[XlsxVisualOccurrence, ...], ...]
    warnings: tuple[WarningRecord, ...]


@dataclass(frozen=True, slots=True)
class XlsxVisualRequest:
    digest: str
    image_path: Path
    prompt: str
    source_index: int
    kind: VisionRequestKind

    def to_vision_request(self) -> VisionRequest:
        return VisionRequest(
            image_path=self.image_path,
            prompt=self.prompt,
            source_index=self.source_index,
            kind=self.kind,
        )


@dataclass(frozen=True, slots=True)
class _OccurrenceWarning:
    code: str
    anchor: str
    detail: str


class _UnsupportedChartType(Exception):
    pass


def _safe_root(archive: ZipFile, part_name: str, *, message: str) -> Any:
    try:
        data = archive.read(part_name)
        return DefusedET.fromstring(
            data,
            forbid_dtd=True,
            forbid_entities=True,
            forbid_external=True,
        )
    except (KeyError, OSError, DefusedXmlException, DefusedET.ParseError) as error:
        raise CorruptDocumentError(message) from error


def _local_name(element: Any) -> str:
    return str(element.tag).rsplit("}", 1)[-1]


def _relationship_index(
    archive: ZipFile,
    infos: dict[str, ZipInfo],
    part_name: str,
) -> dict[str, Any]:
    return _read_relationships(
        archive,
        infos,
        part_name,
        required=_relationships_part(part_name) in infos,
    )


def _relationship(
    relationships: dict[str, Any],
    relationship_id: str | None,
    *,
    expected_type: str,
) -> Any:
    relationship = relationships.get(relationship_id or "")
    if (
        relationship is None
        or relationship.external
        or relationship.relationship_type != expected_type
    ):
        raise CorruptDocumentError("XLSX drawing relationship is invalid")
    return relationship


def _marker_coordinate(marker: Any) -> tuple[int, int]:
    try:
        column = int(marker.findtext(f"{{{_DRAWING_NS}}}col", "")) + 1
        row = int(marker.findtext(f"{{{_DRAWING_NS}}}row", "")) + 1
    except ValueError as error:
        raise CorruptDocumentError("XLSX drawing anchor is invalid") from error
    if not 1 <= column <= 16_384 or not 1 <= row <= 1_048_576:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")
    return row, column


def _drawing_anchor(element: Any) -> str:
    kind = _local_name(element)
    if kind == "absoluteAnchor":
        return "A1"
    start = element.find(f"{{{_DRAWING_NS}}}from")
    if start is None:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")
    start_row, start_column = _marker_coordinate(start)
    start_text = f"{get_column_letter(start_column)}{start_row}"
    if kind == "oneCellAnchor":
        return start_text
    end = element.find(f"{{{_DRAWING_NS}}}to")
    if kind != "twoCellAnchor" or end is None:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")
    end_row, end_column = _marker_coordinate(end)
    if end_row < start_row or end_column < start_column:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")
    end_text = f"{get_column_letter(end_column)}{end_row}"
    return start_text if start_text == end_text else f"{start_text}:{end_text}"


def _metadata(anchor: Any) -> tuple[str | None, str | None, str | None]:
    node = next(anchor.iter(f"{{{_DRAWING_NS}}}cNvPr"), None)
    if node is None:
        return None, None, None
    return node.get("name"), node.get("descr"), node.get("title")


def _plain_text(element: Any | None) -> str:
    if element is None:
        return ""
    rich = "".join(node.text or "" for node in element.iter(f"{{{_DRAWING_MAIN_NS}}}t"))
    if rich:
        return rich.strip()
    values = "".join(node.text or "" for node in element.iter(f"{{{_CHART_NS}}}v"))
    return values.strip()


def _cache_values(reference: Any) -> tuple[str, ...] | None:
    cache = next(
        (
            child
            for child in reference
            if _local_name(child) in {"strCache", "numCache", "multiLvlStrCache"}
        ),
        None,
    )
    if cache is None:
        return None
    if _local_name(cache) == "multiLvlStrCache":
        levels: list[dict[int, str]] = []
        for level in cache.findall(f"{{{_CHART_NS}}}lvl"):
            indexed_level: dict[int, str] = {}
            for point in level.findall(f"{{{_CHART_NS}}}pt"):
                try:
                    index = int(point.get("idx", ""))
                except ValueError as error:
                    raise CorruptDocumentError("XLSX chart cache index is invalid") from error
                if index < 0 or index in indexed_level:
                    raise CorruptDocumentError("XLSX chart cache index is invalid")
                indexed_level[index] = point.findtext(f"{{{_CHART_NS}}}v", "")
            levels.append(indexed_level)
        declared_node = cache.find(f"{{{_CHART_NS}}}ptCount")
        declared = max(
            (max(level, default=-1) + 1 for level in levels),
            default=0,
        )
        if declared_node is not None:
            try:
                declared = int(declared_node.get("val", ""))
            except ValueError as error:
                raise CorruptDocumentError("XLSX chart cache count is invalid") from error
        if declared < 0 or declared > MAX_CHART_CACHE_POINTS:
            raise LimitExceededError("XLSX exceeds the chart cache point limit")
        return tuple(
            " / ".join(value for level in levels if (value := level.get(index, "")))
            for index in range(declared)
        )
    indexed: dict[int, str] = {}
    for point in cache.iter(f"{{{_CHART_NS}}}pt"):
        try:
            index = int(point.get("idx", ""))
        except ValueError as error:
            raise CorruptDocumentError("XLSX chart cache index is invalid") from error
        if index < 0 or index in indexed:
            raise CorruptDocumentError("XLSX chart cache index is invalid")
        indexed[index] = point.findtext(f"{{{_CHART_NS}}}v", "")
    declared_node = next(cache.iter(f"{{{_CHART_NS}}}ptCount"), None)
    declared = max(indexed, default=-1) + 1
    if declared_node is not None:
        try:
            declared = int(declared_node.get("val", ""))
        except ValueError as error:
            raise CorruptDocumentError("XLSX chart cache count is invalid") from error
    if declared < 0 or declared > MAX_CHART_CACHE_POINTS:
        raise LimitExceededError("XLSX exceeds the chart cache point limit")
    return tuple(indexed.get(index, "") for index in range(declared))


def _unquote_sheet(value: str) -> str:
    if value.startswith("'") and value.endswith("'"):
        return value[1:-1].replace("''", "'")
    return value


def _resolve_local_formula(
    formula: str,
    workbook_values: dict[str, dict[str, str]],
) -> tuple[str, ...] | None:
    match = _LOCAL_RANGE.fullmatch(formula.removeprefix("="))
    if match is None:
        return None
    sheet_name = _unquote_sheet(match.group("sheet"))
    values = workbook_values.get(sheet_name)
    if values is None:
        return None
    start = match.group("start").replace("$", "").upper()
    end = (match.group("end") or start).replace("$", "").upper()
    try:
        minimum_column, minimum_row, maximum_column, maximum_row = range_boundaries(
            f"{start}:{end}"
        )
    except ValueError:
        return None
    return tuple(
        values.get(f"{get_column_letter(column)}{row}", "")
        for row in range(minimum_row, maximum_row + 1)
        for column in range(minimum_column, maximum_column + 1)
    )


def _reference_values(
    container: Any | None,
    workbook_values: dict[str, dict[str, str]],
) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    if container is None:
        return (), (), ()
    reference = next(
        (
            node
            for node in container.iter()
            if _local_name(node) in {"strRef", "numRef", "multiLvlStrRef"}
        ),
        None,
    )
    if reference is None:
        direct = container.find(f"{{{_CHART_NS}}}v")
        return ((direct.text or "",), (), ()) if direct is not None else ((), (), ())
    formula = reference.findtext(f"{{{_CHART_NS}}}f", "").strip()
    formulas = (formula,) if formula else ()
    resolved = _resolve_local_formula(formula, workbook_values) if formula else None
    cached = _cache_values(reference)
    if cached is not None:
        return cached, formulas, () if resolved is not None else formulas
    if resolved is not None:
        return resolved, formulas, ()
    return (), formulas, formulas


def _series_facts(
    element: Any,
    *,
    chart_type: str,
    workbook_values: dict[str, dict[str, str]],
) -> XlsxChartSeriesFacts:
    name_values, name_formulas, name_unresolved = _reference_values(
        element.find(f"{{{_CHART_NS}}}tx"),
        workbook_values,
    )
    categories, category_formulas, category_unresolved = _reference_values(
        element.find(f"{{{_CHART_NS}}}cat"),
        workbook_values,
    )
    values, value_formulas, value_unresolved = _reference_values(
        element.find(f"{{{_CHART_NS}}}val"),
        workbook_values,
    )
    x_values, x_formulas, x_unresolved = _reference_values(
        element.find(f"{{{_CHART_NS}}}xVal"),
        workbook_values,
    )
    y_values, y_formulas, y_unresolved = _reference_values(
        element.find(f"{{{_CHART_NS}}}yVal"),
        workbook_values,
    )
    if chart_type == "scatter":
        categories = ()
        values = ()
    formulas = tuple(
        dict.fromkeys(
            (*name_formulas, *category_formulas, *value_formulas, *x_formulas, *y_formulas)
        )
    )
    unresolved = tuple(
        dict.fromkeys(
            (
                *name_unresolved,
                *category_unresolved,
                *value_unresolved,
                *x_unresolved,
                *y_unresolved,
            )
        )
    )
    return XlsxChartSeriesFacts(
        name=next((value for value in name_values if value), "Series"),
        categories=categories,
        values=values,
        x_values=x_values,
        y_values=y_values,
        formulas=formulas,
        unresolved_formulas=unresolved,
    )


def _data_label_facts(chart: Any) -> tuple[str, ...]:
    facts: list[str] = []
    labels = chart.find(f"{{{_CHART_NS}}}dLbls")
    if labels is None:
        return ()
    flag_labels = {
        "showLegendKey": "legend key",
        "showVal": "value",
        "showCatName": "category name",
        "showSerName": "series name",
        "showPercent": "percentage",
        "showBubbleSize": "bubble size",
    }
    for name, label in flag_labels.items():
        node = labels.find(f"{{{_CHART_NS}}}{name}")
        if node is not None and node.get("val", "1") not in {"0", "false", "False"}:
            facts.append(label)
    for item in labels.findall(f"{{{_CHART_NS}}}dLbl"):
        text = _plain_text(item.find(f"{{{_CHART_NS}}}tx"))
        if text:
            facts.append(text)
    return tuple(dict.fromkeys(facts))


def _chart_facts(
    root: Any,
    workbook_values: dict[str, dict[str, str]],
) -> XlsxChartFacts:
    if root.tag != f"{{{_CHART_NS}}}chartSpace":
        raise CorruptDocumentError("XLSX chart namespace is invalid")
    chart = root.find(f"{{{_CHART_NS}}}chart")
    plot_area = chart.find(f"{{{_CHART_NS}}}plotArea") if chart is not None else None
    if chart is None or plot_area is None:
        raise CorruptDocumentError("XLSX chart part is corrupt")
    chart_node = next(
        (child for child in plot_area if _local_name(child) in _CHART_TYPES),
        None,
    )
    if chart_node is None:
        raise _UnsupportedChartType
    chart_type = _CHART_TYPES[_local_name(chart_node)]
    chart_title = chart.find(f"{{{_CHART_NS}}}title")
    title_values, title_formulas, title_unresolved = _reference_values(
        chart_title.find(f"{{{_CHART_NS}}}tx") if chart_title is not None else None,
        workbook_values,
    )
    title = next((value for value in title_values if value), _plain_text(chart_title))
    axis_titles: list[str] = []
    axis_labels: list[str] = []
    axis_formulas: list[str] = []
    axis_unresolved: list[str] = []
    for axis in plot_area:
        if _local_name(axis) not in _AXIS_TYPES:
            continue
        axis_title_node = axis.find(f"{{{_CHART_NS}}}title")
        axis_title_values, formulas, unresolved = _reference_values(
            axis_title_node.find(f"{{{_CHART_NS}}}tx") if axis_title_node is not None else None,
            workbook_values,
        )
        axis_title = next(
            (value for value in axis_title_values if value),
            _plain_text(axis_title_node),
        )
        if axis_title:
            axis_titles.append(axis_title)
        axis_formulas.extend(formulas)
        axis_unresolved.extend(unresolved)
        label_position = axis.find(f"{{{_CHART_NS}}}tickLblPos")
        if label_position is not None and label_position.get("val"):
            axis_labels.append(label_position.get("val", ""))
    return XlsxChartFacts(
        chart_type=chart_type,
        title=title,
        axis_titles=tuple(axis_titles),
        axis_labels=tuple(axis_labels),
        data_labels=_data_label_facts(chart_node),
        formulas=tuple(dict.fromkeys((*title_formulas, *axis_formulas))),
        unresolved_formulas=tuple(dict.fromkeys((*title_unresolved, *axis_unresolved))),
        series=tuple(
            _series_facts(series, chart_type=chart_type, workbook_values=workbook_values)
            for series in chart_node.findall(f"{{{_CHART_NS}}}ser")
        ),
    )


def _paragraph(text: str) -> ParagraphBlock:
    return ParagraphBlock((InlineText(text),))


def chart_fact_blocks(facts: XlsxChartFacts) -> tuple[Block, ...]:
    blocks: list[Block] = [
        HeadingBlock(2, (InlineText(facts.title or f"{facts.chart_type.title()} chart"),)),
        _paragraph(f"Chart type: {facts.chart_type}"),
    ]
    if facts.axis_titles:
        blocks.append(_paragraph(f"Axis titles: {'; '.join(facts.axis_titles)}"))
    if facts.axis_labels:
        blocks.append(_paragraph(f"Axis label positions: {'; '.join(facts.axis_labels)}"))
    if facts.data_labels:
        blocks.append(_paragraph(f"Data labels: {'; '.join(facts.data_labels)}"))
    series_names = tuple(dict.fromkeys(series.name for series in facts.series))
    if series_names:
        blocks.append(_paragraph(f"Series names: {'; '.join(series_names)}"))
    formulas = tuple(
        dict.fromkeys(
            (*facts.formulas, *(formula for item in facts.series for formula in item.formulas))
        )
    )
    if formulas:
        blocks.append(_paragraph(f"Local/formula references: {'; '.join(formulas)}"))
    unresolved = tuple(
        dict.fromkeys(
            (
                *facts.unresolved_formulas,
                *(formula for item in facts.series for formula in item.unresolved_formulas),
            )
        )
    )
    if unresolved:
        blocks.append(
            _paragraph(f"References preserved without evaluation: {'; '.join(unresolved)}")
        )
    rows: list[tuple[str, ...]] = []
    for series in facts.series:
        if facts.chart_type == "scatter":
            width = max(len(series.x_values), len(series.y_values))
            rows.extend(
                (
                    "Series",
                    series.name,
                    "X",
                    series.x_values[index] if index < len(series.x_values) else "",
                    "Y",
                    series.y_values[index] if index < len(series.y_values) else "",
                )
                for index in range(width)
            )
        else:
            width = max(len(series.categories), len(series.values))
            rows.extend(
                (
                    "Series",
                    series.name,
                    "Category",
                    series.categories[index] if index < len(series.categories) else "",
                    "Value",
                    series.values[index] if index < len(series.values) else "",
                )
                for index in range(width)
            )
    if rows:
        blocks.append(TableBlock(tuple(rows), header_rows=0))
    elif facts.series:
        blocks.append(_paragraph("Chart series have no saved cache or resolvable local values."))
    else:
        blocks.append(_paragraph("Chart has no readable series."))
    return tuple(blocks)


def _sample_pairs(
    left: tuple[str, ...],
    right: tuple[str, ...],
) -> tuple[tuple[str, str], ...]:
    count = max(len(left), len(right))
    if count == 0:
        return ()
    if count <= _PREVIEW_SERIES_POINTS:
        indexes = range(count)
    else:
        indexes = tuple(
            round(index * (count - 1) / (_PREVIEW_SERIES_POINTS - 1))
            for index in range(_PREVIEW_SERIES_POINTS)
        )
    return tuple(
        (
            left[index] if index < len(left) else "",
            right[index] if index < len(right) else "",
        )
        for index in indexes
    )


def _preview_lines(facts: XlsxChartFacts) -> tuple[str, ...]:
    lines = [
        "视觉解释输入 / 非 Excel 外观还原",
        f"Chart type: {facts.chart_type}",
        f"Title: {facts.title or '(none)'}",
    ]
    if facts.axis_titles:
        lines.append(f"Axis titles: {'; '.join(facts.axis_titles)}")
    if facts.data_labels:
        lines.append(f"Data labels: {'; '.join(facts.data_labels)}")
    if facts.formulas:
        lines.append(f"Chart references: {'; '.join(facts.formulas)}")
    for series in facts.series:
        lines.append(f"Series: {series.name}")
        if facts.chart_type == "scatter":
            count = max(len(series.x_values), len(series.y_values))
            pairs = _sample_pairs(series.x_values, series.y_values)
            lines.append(
                f"Points (sampled across {count}): " + "; ".join(f"({x}, {y})" for x, y in pairs)
            )
        else:
            count = max(len(series.categories), len(series.values))
            pairs = _sample_pairs(series.categories, series.values)
            lines.append(
                f"Values (sampled across {count}): "
                + "; ".join(f"{category}={value}" for category, value in pairs)
            )
        if series.formulas:
            lines.append(f"References: {'; '.join(series.formulas)}")
    normalized = [line[:_PREVIEW_MAX_LINE_CHARS] for line in lines[:_PREVIEW_MAX_LINES]]
    if len(lines) > _PREVIEW_MAX_LINES:
        normalized.append("(additional native facts omitted from preview only)")
    return tuple(normalized)


def render_chart_semantic_preview(facts: XlsxChartFacts) -> bytes:
    lines = _preview_lines(facts)
    height = _PREVIEW_PADDING * 2 + _PREVIEW_LINE_HEIGHT * len(lines)
    image = Image.new("RGB", (_PREVIEW_WIDTH, max(160, height)), "white")
    try:
        draw = ImageDraw.Draw(image)
        font = ImageFont.load_default(size=18)
        for index, line in enumerate(lines):
            draw.text(
                (_PREVIEW_PADDING, _PREVIEW_PADDING + index * _PREVIEW_LINE_HEIGHT),
                line,
                fill="black",
                font=font,
            )
        output = io.BytesIO()
        image.save(output, format="PNG", optimize=False)
        return output.getvalue()
    finally:
        image.close()


def _write_artifact(directory: Path, name: str, data: bytes) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    root = directory.resolve()
    path = directory / name
    if path.resolve().parent != root:
        raise ValueError("XLSX visual artifact escapes the artifact directory")
    if path.exists():
        if path.read_bytes() != data:
            raise CorruptDocumentError("XLSX visual artifact digest collision")
        return
    path.write_bytes(data)


def _chart_occurrence(
    archive: ZipFile,
    relationship: Any,
    *,
    anchor: str,
    metadata: tuple[str | None, str | None, str | None],
    workbook_values: dict[str, dict[str, str]],
    artifact_dir: Path | None,
    facts_cache: dict[str, XlsxChartFacts],
) -> tuple[XlsxVisualOccurrence, tuple[str, ...]]:
    facts = facts_cache.get(relationship.target)
    if facts is None:
        root = _safe_root(archive, relationship.target, message="XLSX chart part is corrupt")
        facts = _chart_facts(root, workbook_values)
        facts_cache[relationship.target] = facts
    blocks = chart_fact_blocks(facts)
    artifact_name: str | None = None
    digest: str | None = None
    if artifact_dir is not None:
        preview = render_chart_semantic_preview(facts)
        digest = hashlib.sha256(preview).hexdigest()
        artifact_name = f"xlsx-chart-{digest}.png"
        _write_artifact(artifact_dir, artifact_name, preview)
    object_name, alt_text, title = metadata
    warnings = tuple(
        dict.fromkeys(
            (
                *facts.unresolved_formulas,
                *(formula for series in facts.series for formula in series.unresolved_formulas),
            )
        )
    )
    return (
        XlsxVisualOccurrence(
            kind="chart",
            anchor=anchor,
            blocks=blocks,
            artifact_name=artifact_name,
            content_sha256=digest,
            alt_text=alt_text,
            object_name=object_name,
            title=title,
        ),
        warnings,
    )


def _image_occurrence(
    archive: ZipFile,
    relationship: Any,
    *,
    anchor: str,
    metadata: tuple[str | None, str | None, str | None],
    artifact_dir: Path | None,
) -> XlsxVisualOccurrence:
    try:
        data = archive.read(relationship.target)
    except (KeyError, OSError) as error:
        raise CorruptDocumentError("XLSX embedded image part is corrupt") from error
    digest = hashlib.sha256(data).hexdigest()
    suffix = Path(relationship.target).suffix.lower()
    if _SAFE_ARTIFACT_SUFFIX.fullmatch(suffix) is None:
        suffix = ".bin"
    artifact_name = f"xlsx-media-{digest}{suffix}"
    if artifact_dir is not None:
        _write_artifact(artifact_dir, artifact_name, data)
    object_name, alt_text, title = metadata
    description = tuple(
        f"{label}: {value}"
        for label, value in (
            ("Image name", object_name),
            ("Image description", alt_text),
            ("Image title", title),
        )
        if value
    )
    blocks: tuple[Block, ...] = (
        HeadingBlock(3, (InlineText("Embedded image"),)),
        *(_paragraph(value) for value in description),
    )
    return XlsxVisualOccurrence(
        kind="image",
        anchor=anchor,
        blocks=blocks,
        artifact_name=artifact_name if artifact_dir is not None else None,
        content_sha256=digest if artifact_dir is not None else None,
        alt_text=alt_text,
        object_name=object_name,
        title=title,
    )


def _sheet_drawings(
    archive: ZipFile,
    infos: dict[str, ZipInfo],
    sheet: XlsxPreflightSheet,
) -> tuple[str, ...]:
    root = _safe_root(archive, sheet.part_name, message="XLSX sheet part is corrupt")
    relationships = _relationship_index(archive, infos, sheet.part_name)
    targets: list[str] = []
    for node in root.iter(f"{{{_SPREADSHEET_NS}}}drawing"):
        relationship = _relationship(
            relationships,
            node.get(_RELATIONSHIP_ID),
            expected_type=_DRAWING_RELATIONSHIP,
        )
        targets.append(relationship.target)
    return tuple(targets)


def _drawing_occurrences(
    archive: ZipFile,
    infos: dict[str, ZipInfo],
    drawing_part: str,
    *,
    workbook_values: dict[str, dict[str, str]],
    artifact_dir: Path | None,
    chart_facts_cache: dict[str, XlsxChartFacts],
) -> tuple[tuple[XlsxVisualOccurrence, ...], tuple[_OccurrenceWarning, ...]]:
    root = _safe_root(archive, drawing_part, message="XLSX drawing part is corrupt")
    if root.tag != f"{{{_DRAWING_NS}}}wsDr":
        raise CorruptDocumentError("XLSX drawing namespace is invalid")
    relationships = _relationship_index(archive, infos, drawing_part)
    occurrences: list[XlsxVisualOccurrence] = []
    warnings: list[_OccurrenceWarning] = []
    for anchor_node in root:
        if _local_name(anchor_node) not in {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}:
            continue
        anchor = _drawing_anchor(anchor_node)
        metadata = _metadata(anchor_node)
        chart_node = next(anchor_node.iter(f"{{{_CHART_NS}}}chart"), None)
        image_node = next(anchor_node.iter(f"{{{_DRAWING_MAIN_NS}}}blip"), None)
        if chart_node is not None:
            relationship = _relationship(
                relationships,
                chart_node.get(_RELATIONSHIP_ID),
                expected_type=_CHART_RELATIONSHIP,
            )
            try:
                occurrence, formulas = _chart_occurrence(
                    archive,
                    relationship,
                    anchor=anchor,
                    metadata=metadata,
                    workbook_values=workbook_values,
                    artifact_dir=artifact_dir,
                    facts_cache=chart_facts_cache,
                )
            except _UnsupportedChartType:
                warnings.append(
                    _OccurrenceWarning(
                        "xlsx_unsupported_object",
                        anchor,
                        "unsupported chart type was skipped without guessing",
                    )
                )
                continue
            occurrences.append(occurrence)
            warnings.extend(
                _OccurrenceWarning(
                    "xlsx_external_reference",
                    anchor,
                    f"chart reference was preserved without access: {formula}",
                )
                for formula in formulas
            )
        elif image_node is not None:
            relationship = _relationship(
                relationships,
                image_node.get(_RELATIONSHIP_EMBED),
                expected_type=_IMAGE_RELATIONSHIP,
            )
            occurrences.append(
                _image_occurrence(
                    archive,
                    relationship,
                    anchor=anchor,
                    metadata=metadata,
                    artifact_dir=artifact_dir,
                )
            )
    return tuple(occurrences), tuple(warnings)


def _visual_warning(
    code: str,
    sheet: XlsxPreflightSheet,
    anchor: str,
    detail: str,
) -> WarningRecord:
    return WarningRecord(
        code=code,
        message=f"{sheet.name}!{anchor}: {detail}",
    )


def read_xlsx_visual_objects(
    path: Path,
    preflight: XlsxPreflight,
    workbook_values: dict[str, dict[str, str]],
    *,
    artifact_dir: Path | None,
) -> XlsxVisualObjects:
    if preflight.usage.drawing_objects > MAX_DRAWING_OBJECTS:
        raise LimitExceededError("XLSX exceeds the drawing object limit")
    if preflight.usage.chart_cache_points > MAX_CHART_CACHE_POINTS:
        raise LimitExceededError("XLSX exceeds the chart cache point limit")
    by_sheet: list[tuple[XlsxVisualOccurrence, ...]] = []
    warnings: list[WarningRecord] = []
    try:
        with ZipFile(path) as archive:
            infos = {info.filename: info for info in archive.infolist()}
            chart_facts_cache: dict[str, XlsxChartFacts] = {}
            for sheet in preflight.sheets:
                occurrences: list[XlsxVisualOccurrence] = []
                for drawing_part in _sheet_drawings(archive, infos, sheet):
                    drawing_items, drawing_warnings = _drawing_occurrences(
                        archive,
                        infos,
                        drawing_part,
                        workbook_values=workbook_values,
                        artifact_dir=artifact_dir,
                        chart_facts_cache=chart_facts_cache,
                    )
                    occurrences.extend(drawing_items)
                    for warning in drawing_warnings:
                        warnings.append(
                            _visual_warning(
                                warning.code,
                                sheet,
                                warning.anchor,
                                warning.detail,
                            )
                        )
                if artifact_dir is None:
                    warnings.extend(
                        _visual_warning(
                            "xlsx_visual_artifact_unavailable",
                            sheet,
                            occurrence.anchor,
                            f"{occurrence.kind} visual artifact requires an explicit directory",
                        )
                        for occurrence in occurrences
                    )
                by_sheet.append(tuple(occurrences))
    except BadZipFile as error:
        raise CorruptDocumentError("XLSX package is corrupt") from error
    except OSError as error:
        raise CorruptDocumentError("XLSX package could not be read") from error
    return XlsxVisualObjects(tuple(by_sheet), tuple(warnings))


def _artifact_path(artifact_dir: Path, artifact_name: str) -> Path:
    root = artifact_dir.resolve()
    path = artifact_dir / artifact_name
    if path.resolve().parent != root:
        raise ValueError("XLSX visual artifact escapes the artifact directory")
    return path


def build_xlsx_visual_requests(
    document: XlsxDocument,
    artifact_dir: Path,
) -> tuple[XlsxVisualRequest, ...]:
    requests: list[XlsxVisualRequest] = []
    seen: set[tuple[str, str]] = set()
    for sheet in document.sheets:
        for slot in sorted(sheet.slots, key=lambda item: item.source_index):
            if not isinstance(slot, XlsxImageSlot | XlsxChartSlot):
                continue
            slot_kind = "chart" if isinstance(slot, XlsxChartSlot) else "image"
            key = (slot_kind, slot.content_sha256)
            if key in seen:
                continue
            seen.add(key)
            requests.append(
                XlsxVisualRequest(
                    digest=slot.content_sha256,
                    image_path=_artifact_path(artifact_dir, slot.artifact_name),
                    prompt=(
                        XLSX_CHART_VISION_PROMPT
                        if isinstance(slot, XlsxChartSlot)
                        else XLSX_IMAGE_VISION_PROMPT
                    ),
                    source_index=len(requests),
                    kind=VisionRequestKind.PROSE,
                )
            )
    return tuple(requests)


def prepare_xlsx_visual_artifact(
    slot: XlsxImageSlot | XlsxChartSlot,
    artifact_dir: Path,
    output_directory: Path,
    output_stem: str,
) -> PreparedImage:
    return prepare_image(
        _artifact_path(artifact_dir, slot.artifact_name),
        output_directory,
        output_stem,
        slot.artifact_name,
        "embedded",
        None,
    )
