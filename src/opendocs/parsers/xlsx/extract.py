from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

import openpyxl
from defusedxml import ElementTree as DefusedET
from defusedxml.common import DefusedXmlException
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries

from opendocs._models import (
    Block,
    HeadingBlock,
    InlineText,
    MarkdownBlock,
    ParagraphBlock,
    SpannedTableBlock,
    SpannedTableCell,
    TableBlock,
    WarningRecord,
)
from opendocs.errors import CorruptDocumentError, LimitExceededError
from opendocs.parsers.xlsx.models import (
    XlsxDocument,
    XlsxNativeSlot,
    XlsxSheet,
    XlsxSheetKind,
)
from opendocs.parsers.xlsx.preflight import (
    MAX_MATERIALIZED_GRID_CELLS as PREFLIGHT_MAX_MATERIALIZED_GRID_CELLS,
)
from opendocs.parsers.xlsx.preflight import XlsxPreflight, XlsxPreflightSheet
from opendocs.parsers.xlsx.values import format_saved_value

MAX_MATERIALIZED_GRID_CELLS = PREFLIGHT_MAX_MATERIALIZED_GRID_CELLS

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_CELL_TAG = f"{{{_SPREADSHEET_NS}}}c"
_FORMULA_TAG = f"{{{_SPREADSHEET_NS}}}f"
_VALUE_TAG = f"{{{_SPREADSHEET_NS}}}v"
_WARNING_LIMIT_PER_CODE = 20
_EXTERNAL_FORMULA_REFERENCE = re.compile(r"\[[^\]\r\n]+\][^!\r\n]*!")

_Coordinate = tuple[int, int]
_Bounds = tuple[int, int, int, int]


@dataclass(frozen=True, slots=True)
class _FormulaRecord:
    formula_type: str
    text: str | None
    reference: str | None
    attributes: tuple[tuple[str, str], ...]


@dataclass(frozen=True, slots=True)
class _CellRecord:
    coordinate: str
    cell_type: str | None
    formula: _FormulaRecord | None
    cache_present: bool
    cached_value: str | None


@dataclass(frozen=True, slots=True, order=True)
class _WarningEvent:
    code: str
    sheet_index: int
    row: int
    column: int
    sheet_name: str
    detail: str


class _WarningCollector:
    def __init__(self) -> None:
        self._events: set[_WarningEvent] = set()

    def add(
        self,
        code: str,
        *,
        sheet: XlsxPreflightSheet,
        coordinate: str,
        detail: str,
    ) -> None:
        row, column = coordinate_to_tuple(coordinate)
        self._events.add(
            _WarningEvent(
                code=code,
                sheet_index=sheet.sheet_index,
                row=row,
                column=column,
                sheet_name=sheet.name,
                detail=detail,
            )
        )

    def freeze(self) -> tuple[WarningRecord, ...]:
        grouped: dict[str, list[_WarningEvent]] = {}
        for event in sorted(self._events):
            grouped.setdefault(event.code, []).append(event)
        warnings: list[WarningRecord] = []
        for code in sorted(grouped):
            events = grouped[code]
            for event in events[:_WARNING_LIMIT_PER_CODE]:
                coordinate = f"{get_column_letter(event.column)}{event.row}"
                warnings.append(
                    WarningRecord(
                        code=code,
                        message=f"{event.sheet_name}!{coordinate}: {event.detail}",
                    )
                )
            suppressed = len(events) - _WARNING_LIMIT_PER_CODE
            if suppressed > 0:
                warnings.append(
                    WarningRecord(
                        code=code,
                        message=f"{suppressed} additional {code} warnings suppressed",
                    )
                )
        return tuple(warnings)


@dataclass(frozen=True, slots=True)
class _MergeSpec:
    bounds: _Bounds


@dataclass(frozen=True, slots=True)
class _RegionSpec:
    bounds: _Bounds
    header_rows: int
    merges: tuple[_MergeSpec, ...]
    kind: str


@dataclass(slots=True)
class _MaterializationBudget:
    used: int = 0

    def consume(self, cells: int) -> None:
        self.used += cells
        if self.used > MAX_MATERIALIZED_GRID_CELLS:
            raise LimitExceededError("XLSX exceeds the materialized grid limit")


def _safe_worksheet_xml(data: bytes, *, part_name: str) -> Any:
    try:
        root = DefusedET.fromstring(
            data,
            forbid_dtd=True,
            forbid_entities=True,
            forbid_external=True,
        )
    except (DefusedXmlException, DefusedET.ParseError) as error:
        raise CorruptDocumentError(f"XLSX worksheet part is corrupt: {part_name}") from error
    if root.tag != f"{{{_SPREADSHEET_NS}}}worksheet":
        raise CorruptDocumentError(f"XLSX worksheet part is corrupt: {part_name}")
    return root


def _formula_record(element: Any) -> _FormulaRecord:
    attributes = tuple(sorted((str(name), str(value)) for name, value in element.attrib.items()))
    return _FormulaRecord(
        formula_type=element.get("t", "normal"),
        text=element.text if element.text not in {None, ""} else None,
        reference=element.get("ref"),
        attributes=attributes,
    )


def _worksheet_sidecar(data: bytes, *, part_name: str) -> tuple[_CellRecord, ...]:
    root = _safe_worksheet_xml(data, part_name=part_name)
    records: list[_CellRecord] = []
    for element in root.iter(_CELL_TAG):
        coordinate = element.get("r")
        if coordinate is None:
            raise CorruptDocumentError("XLSX cell coordinate is missing")
        formula_node = element.find(_FORMULA_TAG)
        value_node = element.find(_VALUE_TAG)
        records.append(
            _CellRecord(
                coordinate=coordinate,
                cell_type=element.get("t"),
                formula=_formula_record(formula_node) if formula_node is not None else None,
                cache_present=value_node is not None,
                cached_value=value_node.text if value_node is not None else None,
            )
        )
    return tuple(records)


def _read_sidecars(
    path: Path,
    preflight: XlsxPreflight,
) -> dict[int, tuple[_CellRecord, ...]]:
    records: dict[int, tuple[_CellRecord, ...]] = {}
    try:
        with ZipFile(path) as archive:
            for sheet in preflight.sheets:
                if sheet.kind is XlsxSheetKind.CHARTSHEET:
                    records[sheet.sheet_index] = ()
                    continue
                try:
                    data = archive.read(sheet.part_name)
                except KeyError as error:
                    raise CorruptDocumentError("XLSX worksheet part is missing") from error
                records[sheet.sheet_index] = _worksheet_sidecar(
                    data,
                    part_name=sheet.part_name,
                )
    except BadZipFile as error:
        raise CorruptDocumentError("XLSX package is corrupt") from error
    except OSError as error:
        raise CorruptDocumentError("XLSX package could not be read") from error
    return records


def _decode_cached_value(record: _CellRecord) -> object:
    value = record.cached_value
    if value is None:
        return None
    if record.cell_type == "b":
        if value not in {"0", "1"}:
            raise CorruptDocumentError("XLSX formula boolean cache is invalid")
        return value == "1"
    if record.cell_type in {"e", "str", "inlineStr"}:
        return value
    if record.cell_type == "d":
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError as error:
            raise CorruptDocumentError("XLSX formula date cache is invalid") from error
    try:
        return Decimal(value)
    except InvalidOperation as error:
        raise CorruptDocumentError("XLSX formula numeric cache is invalid") from error


def _loaded_formula_text(value: object) -> str | None:
    if isinstance(value, str):
        return value if value.startswith("=") else f"={value}"
    text = getattr(value, "text", None)
    if isinstance(text, str) and text:
        return text if text.startswith("=") else f"={text}"
    return None


def _literal_formula_text(record: _CellRecord, loaded_value: object) -> str | None:
    if record.formula is not None and record.formula.text:
        return (
            record.formula.text
            if record.formula.text.startswith("=")
            else f"={record.formula.text}"
        )
    return _loaded_formula_text(loaded_value)


def _references_external_workbook(record: _CellRecord, loaded_value: object) -> bool:
    formula_text = _literal_formula_text(record, loaded_value)
    return formula_text is not None and _EXTERNAL_FORMULA_REFERENCE.search(formula_text) is not None


def _record_format_warning(
    warning_collector: _WarningCollector,
    *,
    sheet: XlsxPreflightSheet,
    coordinate: str,
    warning: str | None,
) -> None:
    if warning is None:
        return
    warning_collector.add(
        "xlsx_unsupported_number_format",
        sheet=sheet,
        coordinate=coordinate,
        detail=warning,
    )


def _special_formula_text(
    record: _CellRecord,
    loaded_value: object,
    *,
    saved_value: str,
) -> str | None:
    formula = record.formula
    if formula is None:
        return None
    reference = formula.reference or record.coordinate
    if formula.formula_type == "array":
        expression = _literal_formula_text(record, loaded_value) or "(expression unavailable)"
        rendered = f"Array/spill formula {reference}: {expression}"
        if record.cache_present and saved_value:
            rendered += f"; saved value: {saved_value}"
        return rendered
    if formula.formula_type != "dataTable":
        return None
    parameters = [
        f"{name}={value}" for name, value in formula.attributes if name not in {"t", "ref"}
    ]
    suffix = f" ({', '.join(parameters)})" if parameters else ""
    rendered = f"Data-table formula {reference}{suffix}"
    if record.cache_present and saved_value:
        rendered += f"; saved value: {saved_value}"
    return rendered


def _formula_text(
    record: _CellRecord,
    loaded_value: object,
    *,
    number_format: str,
    epoch: datetime,
    conditional_number_format: bool,
    sheet: XlsxPreflightSheet,
    warnings: _WarningCollector,
) -> str:
    if _references_external_workbook(record, loaded_value):
        warnings.add(
            "xlsx_external_reference",
            sheet=sheet,
            coordinate=record.coordinate,
            detail="external workbook formula was preserved without access",
        )
    saved_value = ""
    if record.cache_present:
        formatted = format_saved_value(
            _decode_cached_value(record),
            number_format,
            epoch=epoch,
            conditional_number_format=conditional_number_format,
        )
        saved_value = formatted.text
        _record_format_warning(
            warnings,
            sheet=sheet,
            coordinate=record.coordinate,
            warning=formatted.warning,
        )

    special = _special_formula_text(record, loaded_value, saved_value=saved_value)
    if special is not None:
        if record.formula is not None and record.formula.formula_type == "dataTable":
            warnings.add(
                "xlsx_data_table_formula",
                sheet=sheet,
                coordinate=record.coordinate,
                detail="data-table formula has no portable literal expression",
            )
        if not record.cache_present:
            warnings.add(
                "xlsx_formula_cache_missing",
                sheet=sheet,
                coordinate=record.coordinate,
                detail="formula has no saved cache; formula text was preserved",
            )
        return special
    if record.cache_present:
        return saved_value
    warnings.add(
        "xlsx_formula_cache_missing",
        sheet=sheet,
        coordinate=record.coordinate,
        detail="formula has no saved cache; formula text was preserved",
    )
    return _literal_formula_text(record, loaded_value) or "Formula expression unavailable"


def _conditional_number_format_ranges(worksheet: Any) -> tuple[Any, ...]:
    ranges: list[Any] = []
    for conditional in worksheet.conditional_formatting:
        rules = worksheet.conditional_formatting[conditional]
        changes_number_format = False
        for rule in rules:
            dxf = getattr(rule, "dxf", None)
            if dxf is not None:
                changes_number_format |= getattr(dxf, "numFmt", None) is not None
            elif getattr(rule, "dxfId", None) is not None:
                changes_number_format = True
        if changes_number_format:
            ranges.append(conditional.sqref)
    return tuple(ranges)


def _has_conditional_number_format(coordinate: str, ranges: tuple[Any, ...]) -> bool:
    return any(coordinate in cell_range for cell_range in ranges)


def _cell_texts(
    worksheet: Any,
    records: tuple[_CellRecord, ...],
    *,
    epoch: datetime,
    sheet: XlsxPreflightSheet,
    warnings: _WarningCollector,
) -> tuple[dict[_Coordinate, str], set[_Coordinate]]:
    texts: dict[_Coordinate, str] = {}
    semantic: set[_Coordinate] = set()
    conditional_ranges = _conditional_number_format_ranges(worksheet)
    for record in records:
        cell = worksheet[record.coordinate]
        conditional_number_format = _has_conditional_number_format(
            record.coordinate,
            conditional_ranges,
        )
        if record.formula is not None:
            text = _formula_text(
                record,
                cell.value,
                number_format=cell.number_format,
                epoch=epoch,
                conditional_number_format=conditional_number_format,
                sheet=sheet,
                warnings=warnings,
            )
        else:
            formatted = format_saved_value(
                cell.value,
                cell.number_format,
                epoch=epoch,
                conditional_number_format=conditional_number_format,
            )
            text = formatted.text
            _record_format_warning(
                warnings,
                sheet=sheet,
                coordinate=record.coordinate,
                warning=formatted.warning,
            )
        coordinate = coordinate_to_tuple(record.coordinate)
        texts[coordinate] = text
        if text != "":
            semantic.add(coordinate)
    return texts, semantic


def _area(bounds: _Bounds) -> int:
    minimum_column, minimum_row, maximum_column, maximum_row = bounds
    return (maximum_column - minimum_column + 1) * (maximum_row - minimum_row + 1)


def _coordinates(bounds: _Bounds) -> set[_Coordinate]:
    minimum_column, minimum_row, maximum_column, maximum_row = bounds
    return {
        (row, column)
        for row in range(minimum_row, maximum_row + 1)
        for column in range(minimum_column, maximum_column + 1)
    }


def _anchor(bounds: _Bounds) -> str:
    minimum_column, minimum_row, maximum_column, maximum_row = bounds
    start = f"{get_column_letter(minimum_column)}{minimum_row}"
    end = f"{get_column_letter(maximum_column)}{maximum_row}"
    return start if start == end else f"{start}:{end}"


def _merge_specs(worksheet: Any) -> tuple[_MergeSpec, ...]:
    return tuple(
        sorted(
            (
                _MergeSpec(range_boundaries(str(cell_range)))
                for cell_range in worksheet.merged_cells.ranges
            ),
            key=lambda item: item.bounds,
        )
    )


def _table_specs(worksheet: Any, merges: tuple[_MergeSpec, ...]) -> tuple[_RegionSpec, ...]:
    tables = sorted(worksheet.tables.values(), key=lambda table: range_boundaries(table.ref))
    specs: list[_RegionSpec] = []
    for table in tables:
        bounds = range_boundaries(table.ref)
        table_merges = tuple(merge for merge in merges if _bounds_within(merge.bounds, bounds))
        specs.append(
            _RegionSpec(
                bounds=bounds,
                header_rows=1 if (table.headerRowCount or 0) > 0 else 0,
                merges=table_merges,
                kind="table",
            )
        )
    return tuple(specs)


def _bounds_within(inner: _Bounds, outer: _Bounds) -> bool:
    inner_left, inner_top, inner_right, inner_bottom = inner
    outer_left, outer_top, outer_right, outer_bottom = outer
    return (
        outer_left <= inner_left <= inner_right <= outer_right
        and outer_top <= inner_top <= inner_bottom <= outer_bottom
    )


def _component_specs(
    semantic: set[_Coordinate],
    merges: tuple[_MergeSpec, ...],
    occupied: set[_Coordinate],
) -> tuple[_RegionSpec, ...]:
    available = set(semantic) - occupied
    available_merges: list[_MergeSpec] = []
    for merge in merges:
        footprint = _coordinates(merge.bounds) - occupied
        if not footprint:
            continue
        available.update(footprint)
        available_merges.append(merge)

    remaining = set(available)
    specs: list[_RegionSpec] = []
    for seed in sorted(available):
        if seed not in remaining:
            continue
        stack = [seed]
        remaining.remove(seed)
        component: set[_Coordinate] = set()
        while stack:
            row, column = stack.pop()
            component.add((row, column))
            for neighbor in (
                (row - 1, column),
                (row + 1, column),
                (row, column - 1),
                (row, column + 1),
            ):
                if neighbor in remaining:
                    remaining.remove(neighbor)
                    stack.append(neighbor)
        rows = [row for row, _ in component]
        columns = [column for _, column in component]
        bounds = (min(columns), min(rows), max(columns), max(rows))
        component_merges = tuple(
            merge for merge in available_merges if _bounds_within(merge.bounds, bounds)
        )
        specs.append(_RegionSpec(bounds, 0, component_merges, "region"))
    return tuple(sorted(specs, key=lambda item: item.bounds))


def _region_specs(
    worksheet: Any,
    semantic: set[_Coordinate],
) -> tuple[_RegionSpec, ...]:
    merges = _merge_specs(worksheet)
    table_specs = _table_specs(worksheet, merges)
    occupied: set[_Coordinate] = set()
    for table in table_specs:
        occupied.update(_coordinates(table.bounds))
    component_specs = _component_specs(semantic, merges, occupied)
    return tuple(sorted((*table_specs, *component_specs), key=lambda item: item.bounds))


def _region_block(
    spec: _RegionSpec, texts: dict[_Coordinate, str]
) -> TableBlock | SpannedTableBlock:
    minimum_column, minimum_row, maximum_column, maximum_row = spec.bounds
    row_count = maximum_row - minimum_row + 1
    column_count = maximum_column - minimum_column + 1
    if not spec.merges:
        return TableBlock(
            tuple(
                tuple(
                    texts.get((row, column), "")
                    for column in range(minimum_column, maximum_column + 1)
                )
                for row in range(minimum_row, maximum_row + 1)
            ),
            spec.header_rows,
        )

    merge_by_origin = {(merge.bounds[1], merge.bounds[0]): merge for merge in spec.merges}
    covered: set[_Coordinate] = set()
    for merge in spec.merges:
        covered.update(_coordinates(merge.bounds))
    cells: list[SpannedTableCell] = []
    for row in range(minimum_row, maximum_row + 1):
        for column in range(minimum_column, maximum_column + 1):
            merge = merge_by_origin.get((row, column))
            if merge is not None:
                left, top, right, bottom = merge.bounds
                cells.append(
                    SpannedTableCell(
                        row - minimum_row,
                        column - minimum_column,
                        bottom - top + 1,
                        right - left + 1,
                        texts.get((row, column), ""),
                    )
                )
            elif (row, column) in covered:
                continue
            else:
                cells.append(
                    SpannedTableCell(
                        row - minimum_row,
                        column - minimum_column,
                        1,
                        1,
                        texts.get((row, column), ""),
                    )
                )
    return SpannedTableBlock(row_count, column_count, tuple(cells), spec.header_rows)


def _sheet_prelude(sheet: XlsxPreflightSheet) -> XlsxNativeSlot:
    return XlsxNativeSlot(
        source_index=0,
        anchor="A1",
        blocks=(
            MarkdownBlock(f"<!-- xlsx-sheet: {sheet.sheet_index} -->"),
            HeadingBlock(1, (InlineText(sheet.name),)),
            ParagraphBlock((InlineText(f"Sheet state: {sheet.state.value}"),)),
        ),
    )


def _sheet_slots(
    worksheet: Any,
    records: tuple[_CellRecord, ...],
    *,
    epoch: datetime,
    sheet: XlsxPreflightSheet,
    warnings: _WarningCollector,
    budget: _MaterializationBudget,
) -> tuple[XlsxNativeSlot, ...]:
    texts, semantic = _cell_texts(
        worksheet,
        records,
        epoch=epoch,
        sheet=sheet,
        warnings=warnings,
    )
    slots: list[XlsxNativeSlot] = [_sheet_prelude(sheet)]
    for source_index, spec in enumerate(_region_specs(worksheet, semantic), start=1):
        budget.consume(_area(spec.bounds))
        anchor = _anchor(spec.bounds)
        comment = (
            f"<!-- xlsx-{spec.kind}: sheet={sheet.sheet_index} "
            f"range={anchor} object={source_index} -->"
        )
        block: Block = _region_block(spec, texts)
        slots.append(
            XlsxNativeSlot(
                source_index=source_index,
                anchor=anchor,
                blocks=(MarkdownBlock(comment), block),
            )
        )
    return tuple(slots)


def extract_xlsx(path: Path, preflight: XlsxPreflight) -> XlsxDocument:
    if not isinstance(path, Path):
        raise TypeError("path must be a Path")
    if not isinstance(preflight, XlsxPreflight):
        raise TypeError("preflight must be an XlsxPreflight")
    sidecars = _read_sidecars(path, preflight)
    warnings = _WarningCollector()
    workbook = openpyxl.load_workbook(
        path,
        read_only=False,
        data_only=False,
        rich_text=False,
        keep_links=False,
    )
    try:
        worksheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}
        sheets: list[XlsxSheet] = []
        budget = _MaterializationBudget()
        for sheet in preflight.sheets:
            if sheet.kind is XlsxSheetKind.CHARTSHEET:
                slots = (_sheet_prelude(sheet),)
            else:
                worksheet = worksheets.get(sheet.name)
                if worksheet is None:
                    raise CorruptDocumentError("XLSX worksheet is missing after full-mode load")
                slots = _sheet_slots(
                    worksheet,
                    sidecars[sheet.sheet_index],
                    epoch=workbook.epoch,
                    sheet=sheet,
                    warnings=warnings,
                    budget=budget,
                )
            sheets.append(
                XlsxSheet(
                    sheet_index=sheet.sheet_index,
                    name=sheet.name,
                    kind=sheet.kind,
                    state=sheet.state,
                    slots=slots,
                )
            )
    finally:
        workbook.close()
    return XlsxDocument(sheets=tuple(sheets), warnings=warnings.freeze())
