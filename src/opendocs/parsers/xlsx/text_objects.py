from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlsplit
from zipfile import BadZipFile, ZipFile, ZipInfo

from defusedxml import ElementTree as DefusedET
from defusedxml.common import DefusedXmlException
from openpyxl.utils.cell import get_column_letter, range_boundaries

from opendocs._models import Block, InlineLink, InlineText, MarkdownBlock, ParagraphBlock
from opendocs.errors import CorruptDocumentError
from opendocs.parsers.xlsx.preflight import (
    XlsxPreflight,
    XlsxPreflightSheet,
    _read_relationships,
    _relationships_part,
)

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_THREADED_REL_NS = "http://schemas.microsoft.com/office/2017/10/relationships"
_THREADED_COMMENTS_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"

_COMMENTS_RELATIONSHIP = f"{_OFFICE_REL_NS}/comments"
_DRAWING_RELATIONSHIP = f"{_OFFICE_REL_NS}/drawing"
_HYPERLINK_RELATIONSHIP = f"{_OFFICE_REL_NS}/hyperlink"
_EXTERNAL_LINK_RELATIONSHIP = f"{_OFFICE_REL_NS}/externalLink"
_EXTERNAL_LINK_PATH_RELATIONSHIP = f"{_OFFICE_REL_NS}/externalLinkPath"
_CONNECTIONS_RELATIONSHIP = f"{_OFFICE_REL_NS}/connections"
_THREADED_COMMENTS_RELATIONSHIP = f"{_THREADED_REL_NS}/threadedComment"
_PERSON_RELATIONSHIP = f"{_THREADED_REL_NS}/person"
_RELATIONSHIP_ID = f"{{{_OFFICE_REL_NS}}}id"

_KIND_HYPERLINK = 10
_KIND_COMMENT = 20
_KIND_TEXT_BOX = 30
_KIND_HEADER_FOOTER = 40
_KIND_EXTERNAL_REFERENCE = 50
_SAFE_LINK_SCHEMES = {"http", "https", "mailto"}
_UNSUPPORTED_KIND = re.compile(r"[^A-Za-z0-9_.-]+")
_HEADER_TAGS = (
    ("oddHeader", "Odd header"),
    ("oddFooter", "Odd footer"),
    ("evenHeader", "Even header"),
    ("evenFooter", "Even footer"),
    ("firstHeader", "First header"),
    ("firstFooter", "First footer"),
)
_HEADER_SECTIONS = (("L", "left"), ("C", "center"), ("R", "right"))
_HEADER_FIELDS = {
    "P": "{page}",
    "N": "{pages}",
    "D": "{date}",
    "T": "{time}",
    "F": "{file}",
    "Z": "{path}",
    "A": "{sheet}",
}
_HEADER_NAMED_FIELDS = {
    "Page": "{page}",
    "Pages": "{pages}",
    "Date": "{date}",
    "Time": "{time}",
    "File": "{file}",
    "Path": "{path}",
    "Tab": "{sheet}",
}
_HEADER_FORMAT_CODES = {"B", "I", "U", "E", "S", "X", "Y", "O", "H"}


@dataclass(frozen=True, slots=True)
class XlsxTextObject:
    sheet_index: int
    anchor: str
    row: int
    column: int
    kind_rank: int
    source_ordinal: int
    paragraphs: tuple[str, ...] = ()
    link_label: str | None = None
    link_target: str | None = None
    safe_link: bool = False


@dataclass(frozen=True, slots=True)
class XlsxTextWarning:
    code: str
    sheet_index: int
    anchor: str
    object_ordinal: int
    detail: str


@dataclass(frozen=True, slots=True)
class XlsxTextObjects:
    by_sheet: tuple[tuple[XlsxTextObject, ...], ...]
    warnings: tuple[XlsxTextWarning, ...]


@dataclass(slots=True)
class _SheetCollector:
    sheet: XlsxPreflightSheet
    objects: list[XlsxTextObject]
    warnings: list[XlsxTextWarning]
    next_ordinal: int = 1

    def add_object(
        self,
        *,
        anchor: str,
        kind_rank: int,
        paragraphs: tuple[str, ...] = (),
        link_label: str | None = None,
        link_target: str | None = None,
        safe_link: bool = False,
    ) -> int:
        row, column = _top_left(anchor)
        ordinal = self.next_ordinal
        self.next_ordinal += 1
        self.objects.append(
            XlsxTextObject(
                sheet_index=self.sheet.sheet_index,
                anchor=anchor,
                row=row,
                column=column,
                kind_rank=kind_rank,
                source_ordinal=ordinal,
                paragraphs=paragraphs,
                link_label=link_label,
                link_target=link_target,
                safe_link=safe_link,
            )
        )
        return ordinal

    def add_warning(
        self,
        code: str,
        *,
        anchor: str,
        detail: str,
        object_ordinal: int | None = None,
    ) -> None:
        ordinal = object_ordinal
        if ordinal is None:
            ordinal = self.next_ordinal
            self.next_ordinal += 1
        else:
            self.next_ordinal = max(self.next_ordinal, ordinal + 1)
        self.warnings.append(
            XlsxTextWarning(
                code=code,
                sheet_index=self.sheet.sheet_index,
                anchor=anchor,
                object_ordinal=ordinal,
                detail=detail,
            )
        )


def _safe_root(archive: ZipFile, part_name: str, *, message: str) -> Any:
    try:
        data = archive.read(part_name)
        return DefusedET.fromstring(
            data,
            forbid_dtd=True,
            forbid_entities=True,
            forbid_external=True,
        )
    except (KeyError, OSError, DefusedXmlException, DefusedET.ParseError) as error:
        raise CorruptDocumentError(message) from error


def _top_left(anchor: str) -> tuple[int, int]:
    try:
        minimum_column, minimum_row, _, _ = range_boundaries(anchor)
    except ValueError as error:
        raise CorruptDocumentError("XLSX object anchor is invalid") from error
    return minimum_row, minimum_column


def _relationship_index(
    archive: ZipFile,
    infos: dict[str, ZipInfo],
    part_name: str,
) -> dict[str, Any]:
    return _read_relationships(
        archive,
        infos,
        part_name,
        required=_relationships_part(part_name) in infos,
    )


def _relationship(
    relationships: dict[str, Any],
    relationship_id: str | None,
    *,
    expected_type: str,
) -> Any:
    relationship = relationships.get(relationship_id or "")
    if relationship is None or relationship.relationship_type != expected_type:
        raise CorruptDocumentError("XLSX object relationship is invalid")
    return relationship


def _person_names(
    archive: ZipFile,
    infos: dict[str, ZipInfo],
) -> dict[str, str]:
    relationships = _relationship_index(archive, infos, "xl/workbook.xml")
    person_targets = [
        relationship.target
        for relationship in relationships.values()
        if relationship.relationship_type == _PERSON_RELATIONSHIP and not relationship.external
    ]
    if (
        any(
            relationship.external
            for relationship in relationships.values()
            if relationship.relationship_type == _PERSON_RELATIONSHIP
        )
        or len(person_targets) > 1
    ):
        raise CorruptDocumentError("XLSX persons relationship is invalid")
    if not person_targets:
        return {}
    root = _safe_root(archive, person_targets[0], message="XLSX persons part is corrupt")
    if root.tag != f"{{{_THREADED_COMMENTS_NS}}}personList":
        raise CorruptDocumentError("XLSX persons namespace is invalid")
    people: dict[str, str] = {}
    for person in root.findall(f"{{{_THREADED_COMMENTS_NS}}}person"):
        person_id = person.get("id")
        display_name = person.get("displayName")
        if not person_id or not display_name or person_id in people:
            raise CorruptDocumentError("XLSX person entry is invalid")
        people[person_id] = display_name
    return people


def _classic_comments(
    archive: ZipFile,
    relationship: Any,
    collector: _SheetCollector,
) -> None:
    if relationship.external:
        raise CorruptDocumentError("XLSX comments relationship is invalid")
    root = _safe_root(archive, relationship.target, message="XLSX comments part is corrupt")
    if root.tag != f"{{{_SPREADSHEET_NS}}}comments":
        raise CorruptDocumentError("XLSX comments namespace is invalid")
    authors_node = root.find(f"{{{_SPREADSHEET_NS}}}authors")
    authors = (
        tuple(node.text or "" for node in authors_node.findall(f"{{{_SPREADSHEET_NS}}}author"))
        if authors_node is not None
        else ()
    )
    comment_list = root.find(f"{{{_SPREADSHEET_NS}}}commentList")
    if comment_list is None:
        raise CorruptDocumentError("XLSX comments part is corrupt")
    for comment in comment_list.findall(f"{{{_SPREADSHEET_NS}}}comment"):
        reference = comment.get("ref")
        try:
            author_index = int(comment.get("authorId", ""))
        except ValueError as error:
            raise CorruptDocumentError("XLSX comment author is invalid") from error
        if reference is None or not 0 <= author_index < len(authors):
            raise CorruptDocumentError("XLSX comment author or anchor is invalid")
        _top_left(reference)
        text = "".join(node.text or "" for node in comment.iter(f"{{{_SPREADSHEET_NS}}}t"))
        author = authors[author_index]
        prefix = f"Comment by {author}: " if author else "Comment: "
        collector.add_object(
            anchor=reference,
            kind_rank=_KIND_COMMENT,
            paragraphs=(f"{prefix}{text}",),
        )


def _threaded_comments(
    archive: ZipFile,
    relationship: Any,
    people: dict[str, str],
    collector: _SheetCollector,
) -> None:
    if relationship.external:
        raise CorruptDocumentError("XLSX threaded comments relationship is invalid")
    root = _safe_root(
        archive,
        relationship.target,
        message="XLSX threaded comments part is corrupt",
    )
    if root.tag != f"{{{_THREADED_COMMENTS_NS}}}ThreadedComments":
        raise CorruptDocumentError("XLSX threaded comments namespace is invalid")
    for comment in root.findall(f"{{{_THREADED_COMMENTS_NS}}}threadedComment"):
        reference = comment.get("ref")
        person_id = comment.get("personId")
        if reference is None or not person_id or person_id not in people:
            raise CorruptDocumentError("XLSX threaded comment person or anchor is invalid")
        _top_left(reference)
        text = "".join(node.text or "" for node in comment.iter(f"{{{_THREADED_COMMENTS_NS}}}text"))
        collector.add_object(
            anchor=reference,
            kind_rank=_KIND_COMMENT,
            paragraphs=(f"Threaded comment by {people[person_id]}: {text}",),
        )


def _safe_link_target(target: str) -> bool:
    if target.startswith("#"):
        return not any(character.isspace() for character in target)
    if any(character.isspace() or ord(character) < 32 for character in target):
        return False
    try:
        return urlsplit(target).scheme.lower() in _SAFE_LINK_SCHEMES
    except ValueError:
        return False


def _internal_target(location: str) -> str:
    return f"#{_quoted_location(location)}"


def _quoted_location(location: str) -> str:
    return quote(location, safe="!$&'()*+,-./:;=@_~?")


def _hyperlinks(
    worksheet_root: Any,
    relationships: dict[str, Any],
    collector: _SheetCollector,
) -> None:
    for hyperlink in worksheet_root.iter(f"{{{_SPREADSHEET_NS}}}hyperlink"):
        reference = hyperlink.get("ref")
        if reference is None:
            raise CorruptDocumentError("XLSX hyperlink anchor is invalid")
        _top_left(reference)
        relationship_id = hyperlink.get(_RELATIONSHIP_ID)
        location = hyperlink.get("location") or ""
        target = ""
        if relationship_id is not None:
            target = _relationship(
                relationships,
                relationship_id,
                expected_type=_HYPERLINK_RELATIONSHIP,
            ).target
        if target and location:
            target = f"{target}#{_quoted_location(location)}"
        elif location:
            target = location if location.startswith("[") else _internal_target(location)
        if not target:
            raise CorruptDocumentError("XLSX hyperlink target is missing")
        safe = _safe_link_target(target)
        ordinal = collector.add_object(
            anchor=reference,
            kind_rank=_KIND_HYPERLINK,
            link_label=hyperlink.get("display"),
            link_target=target,
            safe_link=safe,
        )
        if not safe:
            collector.add_warning(
                "xlsx_external_reference",
                anchor=reference,
                object_ordinal=ordinal,
                detail="hyperlink target was preserved as plain text without access",
            )


def _drawing_anchor(anchor: Any) -> str:
    local_name = anchor.tag.rsplit("}", 1)[-1]
    if local_name == "absoluteAnchor":
        return "A1"
    start = anchor.find(f"{{{_DRAWING_NS}}}from")
    if start is None:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")

    def coordinate(marker: Any) -> tuple[int, int]:
        try:
            column = int(marker.findtext(f"{{{_DRAWING_NS}}}col", "")) + 1
            row = int(marker.findtext(f"{{{_DRAWING_NS}}}row", "")) + 1
        except ValueError as error:
            raise CorruptDocumentError("XLSX drawing anchor is invalid") from error
        if not 1 <= column <= 16_384 or not 1 <= row <= 1_048_576:
            raise CorruptDocumentError("XLSX drawing anchor is invalid")
        return row, column

    start_row, start_column = coordinate(start)
    start_address = f"{get_column_letter(start_column)}{start_row}"
    if local_name == "oneCellAnchor":
        return start_address
    end = anchor.find(f"{{{_DRAWING_NS}}}to")
    if end is None:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")
    end_row, end_column = coordinate(end)
    if end_row < start_row or end_column < start_column:
        raise CorruptDocumentError("XLSX drawing anchor is invalid")
    end_address = f"{get_column_letter(end_column)}{end_row}"
    return start_address if start_address == end_address else f"{start_address}:{end_address}"


def _drawing_text_boxes(
    archive: ZipFile,
    relationship: Any,
    collector: _SheetCollector,
) -> None:
    if relationship.external:
        raise CorruptDocumentError("XLSX drawing relationship is invalid")
    root = _safe_root(archive, relationship.target, message="XLSX drawing part is corrupt")
    if root.tag != f"{{{_DRAWING_NS}}}wsDr":
        raise CorruptDocumentError("XLSX drawing namespace is invalid")
    for anchor in root:
        if anchor.tag not in {
            f"{{{_DRAWING_NS}}}oneCellAnchor",
            f"{{{_DRAWING_NS}}}twoCellAnchor",
            f"{{{_DRAWING_NS}}}absoluteAnchor",
        }:
            continue
        address = _drawing_anchor(anchor)
        for shape in anchor.iter(f"{{{_DRAWING_NS}}}sp"):
            paragraphs: list[str] = []
            text_paragraphs = [
                "".join(node.text or "" for node in paragraph.iter(f"{{{_DRAWING_MAIN_NS}}}t"))
                for paragraph in shape.iter(f"{{{_DRAWING_MAIN_NS}}}p")
            ]
            text = "\n".join(value for value in text_paragraphs if value)
            if text:
                paragraphs.append(f"Text box: {text}")
            metadata = shape.find(f".//{{{_DRAWING_NS}}}cNvPr")
            if metadata is not None:
                for attribute, label in (
                    ("name", "Shape name"),
                    ("descr", "Alt text"),
                    ("title", "Alt title"),
                ):
                    value = metadata.get(attribute)
                    if value:
                        paragraphs.append(f"{label}: {value}")
            if paragraphs:
                collector.add_object(
                    anchor=address,
                    kind_rank=_KIND_TEXT_BOX,
                    paragraphs=tuple(paragraphs),
                )


def _split_header_footer(raw: str) -> tuple[dict[str, str], set[str]]:
    sections = {"L": [], "C": [], "R": []}
    image_sections: set[str] = set()
    current = "C"
    index = 0
    while index < len(raw):
        if raw[index] != "&":
            sections[current].append(raw[index])
            index += 1
            continue
        if index + 1 >= len(raw):
            sections[current].append("&")
            break
        code = raw[index + 1]
        index += 2
        if code == "&":
            sections[current].append("&")
        elif code in sections:
            current = code
        elif code == '"':
            closing = raw.find('"', index)
            index = len(raw) if closing < 0 else closing + 1
        elif code == "K":
            index = min(index + 6, len(raw))
        elif code == "[":
            closing = raw.find("]", index)
            if closing < 0:
                sections[current].append("&[")
                continue
            field = raw[index:closing]
            index = closing + 1
            if field == "Picture":
                image_sections.add(current)
            elif field in _HEADER_NAMED_FIELDS:
                sections[current].append(_HEADER_NAMED_FIELDS[field])
            else:
                sections[current].append(f"&[{field}]")
        elif code.isdigit():
            while index < len(raw) and raw[index].isdigit():
                index += 1
        elif code in _HEADER_FIELDS:
            sections[current].append(_HEADER_FIELDS[code])
        elif code == "G":
            image_sections.add(current)
        elif code in _HEADER_FORMAT_CODES:
            continue
        else:
            sections[current].append(f"&{code}")
    return (
        {section: "".join(value).strip() for section, value in sections.items()},
        image_sections,
    )


def _headers_and_footers(worksheet_root: Any, collector: _SheetCollector) -> None:
    for tag, label in _HEADER_TAGS:
        element = worksheet_root.find(f".//{{{_SPREADSHEET_NS}}}{tag}")
        if element is None or element.text is None:
            continue
        sections, image_sections = _split_header_footer(element.text)
        for section, section_label in _HEADER_SECTIONS:
            value = sections[section]
            ordinal: int | None = None
            if value:
                ordinal = collector.add_object(
                    anchor="A1",
                    kind_rank=_KIND_HEADER_FOOTER,
                    paragraphs=(f"{label} {section_label}: {value}",),
                )
            if section in image_sections:
                collector.add_warning(
                    "xlsx_unsupported_object",
                    anchor="A1",
                    object_ordinal=ordinal,
                    detail="header/footer image field was skipped",
                )


def _visible_vml_text(archive: ZipFile, part_name: str) -> bool:
    root = _safe_root(archive, part_name, message="XLSX VML drawing part is corrupt")
    return any(
        "".join(node.itertext()).strip()
        for node in root.iter()
        if isinstance(node.tag, str) and node.tag.rsplit("}", 1)[-1] == "textbox"
    )


def _unsupported_objects(
    archive: ZipFile,
    worksheet_root: Any,
    collector: _SheetCollector,
) -> None:
    for reference in collector.sheet.unsupported_objects:
        if reference.kind == "vmlDrawing" and not _visible_vml_text(archive, reference.target):
            continue
        kind = _UNSUPPORTED_KIND.sub("_", reference.kind).strip("_") or "unknown"
        collector.add_warning(
            "xlsx_unsupported_object",
            anchor="A1",
            object_ordinal=reference.source_index + 1,
            detail=f"unsupported {kind} object was skipped",
        )
    for _extension in worksheet_root.iter(f"{{{_SPREADSHEET_NS}}}ext"):
        collector.add_warning(
            "xlsx_unsupported_object",
            anchor="A1",
            detail="unsupported vendor extension was skipped",
        )


def _external_reference(
    collector: _SheetCollector,
    target: str,
    *,
    detail: str,
) -> None:
    ordinal = collector.add_object(
        anchor="A1",
        kind_rank=_KIND_EXTERNAL_REFERENCE,
        paragraphs=(f"External reference: {target}",),
    )
    collector.add_warning(
        "xlsx_external_reference",
        anchor="A1",
        object_ordinal=ordinal,
        detail=detail,
    )


def _workbook_external_references(
    archive: ZipFile,
    infos: dict[str, ZipInfo],
    collector: _SheetCollector,
) -> None:
    relationships = _relationship_index(archive, infos, "xl/workbook.xml")
    for relationship in relationships.values():
        if relationship.relationship_type == _EXTERNAL_LINK_RELATIONSHIP:
            if relationship.external:
                _external_reference(
                    collector,
                    relationship.target,
                    detail="external workbook reference was preserved without access",
                )
                continue
            nested = _relationship_index(archive, infos, relationship.target)
            for nested_relationship in nested.values():
                if nested_relationship.relationship_type == _EXTERNAL_LINK_PATH_RELATIONSHIP:
                    _external_reference(
                        collector,
                        nested_relationship.target,
                        detail="external workbook reference was preserved without access",
                    )
        elif relationship.relationship_type == _CONNECTIONS_RELATIONSHIP:
            if relationship.external:
                raise CorruptDocumentError("XLSX connections relationship is invalid")
            root = _safe_root(
                archive,
                relationship.target,
                message="XLSX connections part is corrupt",
            )
            if root.tag != f"{{{_SPREADSHEET_NS}}}connections":
                raise CorruptDocumentError("XLSX connections namespace is invalid")
            for connection in root.findall(f"{{{_SPREADSHEET_NS}}}connection"):
                for attribute in ("sourceFile", "odcFile", "connectionFile"):
                    target = connection.get(attribute)
                    if target:
                        _external_reference(
                            collector,
                            target,
                            detail="external data reference was preserved without access",
                        )


def read_xlsx_text_objects(path: Path, preflight: XlsxPreflight) -> XlsxTextObjects:
    collectors = {
        sheet.sheet_index: _SheetCollector(sheet=sheet, objects=[], warnings=[])
        for sheet in preflight.sheets
    }
    try:
        with ZipFile(path) as archive:
            infos = {info.filename: info for info in archive.infolist()}
            people = _person_names(archive, infos)
            for sheet in preflight.sheets:
                collector = collectors[sheet.sheet_index]
                worksheet_root = _safe_root(
                    archive,
                    sheet.part_name,
                    message="XLSX worksheet part is corrupt",
                )
                if sheet.kind.value != "worksheet":
                    _unsupported_objects(archive, worksheet_root, collector)
                    continue
                if worksheet_root.tag != f"{{{_SPREADSHEET_NS}}}worksheet":
                    raise CorruptDocumentError("XLSX worksheet namespace is invalid")
                relationships = _relationship_index(archive, infos, sheet.part_name)
                _hyperlinks(worksheet_root, relationships, collector)
                for relationship in relationships.values():
                    if relationship.relationship_type == _COMMENTS_RELATIONSHIP:
                        _classic_comments(archive, relationship, collector)
                    elif relationship.relationship_type == _THREADED_COMMENTS_RELATIONSHIP:
                        _threaded_comments(archive, relationship, people, collector)
                for drawing in worksheet_root.iter(f"{{{_SPREADSHEET_NS}}}drawing"):
                    relationship = _relationship(
                        relationships,
                        drawing.get(_RELATIONSHIP_ID),
                        expected_type=_DRAWING_RELATIONSHIP,
                    )
                    _drawing_text_boxes(archive, relationship, collector)
                _headers_and_footers(worksheet_root, collector)
                _unsupported_objects(archive, worksheet_root, collector)
            if preflight.sheets:
                _workbook_external_references(archive, infos, collectors[1])
    except BadZipFile as error:
        raise CorruptDocumentError("XLSX package is corrupt") from error
    except OSError as error:
        raise CorruptDocumentError("XLSX package could not be read") from error
    warnings = tuple(
        warning for sheet in preflight.sheets for warning in collectors[sheet.sheet_index].warnings
    )
    return XlsxTextObjects(
        by_sheet=tuple(tuple(collectors[sheet.sheet_index].objects) for sheet in preflight.sheets),
        warnings=warnings,
    )


def text_object_blocks(
    item: XlsxTextObject,
    *,
    fallback_label: str,
    object_index: int,
) -> tuple[Block, ...]:
    marker = MarkdownBlock(
        f"<!-- xlsx-object: sheet={item.sheet_index} range={item.anchor} object={object_index} -->"
    )
    if item.link_target is not None:
        label = item.link_label or fallback_label or item.link_target
        inline = (
            InlineLink(label, item.link_target)
            if item.safe_link
            else InlineText(label if label == item.link_target else f"{label} ({item.link_target})")
        )
        return marker, ParagraphBlock((inline,))
    return (marker, *(ParagraphBlock((InlineText(text),)) for text in item.paragraphs))
