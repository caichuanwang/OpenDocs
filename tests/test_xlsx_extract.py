from __future__ import annotations

import hashlib
import socket
import urllib.request
from copy import deepcopy
from dataclasses import replace
from pathlib import Path
from typing import Any, cast
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.numbers import NumberFormat
from openpyxl.worksheet.table import Table
from PIL import Image as PILImage

import opendocs.parsers.xlsx.extract as extract_module
import opendocs.parsers.xlsx.media as media_module
from opendocs._models import (
    DocumentType,
    HeadingBlock,
    InlineLink,
    InlineText,
    MarkdownBlock,
    ParagraphBlock,
    ParsedDocument,
    SpannedTableBlock,
    TableBlock,
)
from opendocs.errors import CorruptDocumentError, DocumentTypeMismatchError, LimitExceededError
from opendocs.markdown import render_markdown
from opendocs.parsers.xlsx.extract import extract_xlsx
from opendocs.parsers.xlsx.media import (
    XLSX_CHART_VISION_PROMPT,
    build_xlsx_visual_requests,
    prepare_xlsx_visual_artifact,
)
from opendocs.parsers.xlsx.models import (
    XlsxChartSlot,
    XlsxImageSlot,
    XlsxNativeSlot,
    XlsxSheet,
    XlsxSheetKind,
    XlsxSheetState,
)
from opendocs.parsers.xlsx.preflight import preflight_xlsx
from tests.xlsx_fixtures import rewrite_xlsx

SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
THREADED_REL_NS = "http://schemas.microsoft.com/office/2017/10/relationships"
THREADED_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"


def _append_before(xml: bytes, closing_tag: bytes, body: str) -> bytes:
    assert closing_tag in xml
    return xml.replace(closing_tag, body.encode() + closing_tag)


def _with_relationship_namespace(xml: bytes) -> bytes:
    if b"xmlns:r=" in xml:
        return xml
    return xml.replace(
        b"<worksheet ",
        f'<worksheet xmlns:r="{OFFICE_REL_NS}" '.encode(),
        1,
    )


def _relationships(*items: tuple[str, str, str, bool]) -> bytes:
    values = "".join(
        f'<Relationship Id="{relationship_id}" Type="{relationship_type}" '
        f'Target="{target}"{_target_mode(external)}/>'
        for relationship_id, relationship_type, target, external in items
    )
    return f'<Relationships xmlns="{PACKAGE_REL_NS}">{values}</Relationships>'.encode()


def _target_mode(external: bool) -> str:
    return ' TargetMode="External"' if external else ""


def _paragraph_text(slot: XlsxNativeSlot) -> str:
    return "\n".join(
        "".join(
            inline.label if isinstance(inline, InlineLink) else inline.text
            for inline in block.inlines
        )
        for block in slot.blocks
        if isinstance(block, ParagraphBlock)
    )


def _save_ordered_workbook(path: Path) -> None:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "Visible"
    visible["A1"] = "kept"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    very_hidden = workbook.create_sheet("Very Hidden")
    very_hidden.sheet_state = "veryHidden"
    chart = BarChart()
    chart.add_data(Reference(visible, min_col=1, min_row=1, max_row=1))
    chart_sheet = workbook.create_chartsheet("Chart")
    chart_sheet.add_chart(chart)
    workbook.save(path)


def _native_slots(document_sheet: XlsxSheet) -> tuple[XlsxNativeSlot, ...]:
    return tuple(slot for slot in document_sheet.slots if isinstance(slot, XlsxNativeSlot))


def _visual_slots(document_sheet: XlsxSheet) -> tuple[XlsxImageSlot | XlsxChartSlot, ...]:
    return tuple(
        slot for slot in document_sheet.slots if isinstance(slot, XlsxImageSlot | XlsxChartSlot)
    )


def _write_png(path: Path, *, color: str = "navy") -> bytes:
    image = PILImage.new("RGB", (32, 16), color)
    try:
        image.save(path, format="PNG")
    finally:
        image.close()
    return path.read_bytes()


def test_extracts_native_chart_facts_and_raw_image_artifacts(tmp_path: Path) -> None:
    path = tmp_path / "visuals.xlsx"
    source_image = tmp_path / "source.png"
    image_bytes = _write_png(source_image)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Data"
    for row in (("Month", "Revenue"), ("Jan", 10), ("Feb", 15), ("Mar", 12)):
        sheet.append(row)
    chart = LineChart()
    chart.title = "Revenue trend"
    cast(Any, chart.x_axis).title = "Month"
    cast(Any, chart.y_axis).title = "USD"
    chart.dataLabels = DataLabelList(showVal=True)
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=4))
    sheet.add_chart(chart, "D2")
    embedded = Image(source_image)
    sheet.add_image(embedded, "K3")
    workbook.save(path)
    with ZipFile(path) as archive:
        drawing_xml = archive.read("xl/drawings/drawing1.xml")
    assert b'name="Chart 1"' in drawing_xml
    assert b'name="Image 2" descr="Picture"' in drawing_xml
    drawing_xml = drawing_xml.replace(
        b'name="Chart 1"',
        b'name="Chart 1" descr="chart description" title="chart title"',
    ).replace(
        b'name="Image 2" descr="Picture"',
        b'name="Image 2" descr="image description" title="image title"',
    )
    rewrite_xlsx(path, {"xl/drawings/drawing1.xml": drawing_xml})
    artifacts = tmp_path / "artifacts"

    document = extract_xlsx(path, preflight_xlsx(path), artifact_dir=artifacts)

    slots = _visual_slots(document.sheets[0])
    chart_slot = next(slot for slot in slots if isinstance(slot, XlsxChartSlot))
    image_slot = next(slot for slot in slots if isinstance(slot, XlsxImageSlot))
    assert chart_slot.anchor == "D2"
    assert image_slot.anchor == "K3"
    assert image_slot.content_sha256 == hashlib.sha256(image_bytes).hexdigest()
    assert (chart_slot.object_name, chart_slot.alt_text, chart_slot.title) == (
        "Chart 1",
        "chart description",
        "chart title",
    )
    assert (image_slot.object_name, image_slot.alt_text, image_slot.title) == (
        "Image 2",
        "image description",
        "image title",
    )
    assert (artifacts / image_slot.artifact_name).read_bytes() == image_bytes
    chart_bytes = (artifacts / chart_slot.artifact_name).read_bytes()
    assert chart_bytes.startswith(b"\x89PNG\r\n\x1a\n")
    assert chart_slot.content_sha256 == hashlib.sha256(chart_bytes).hexdigest()
    headings = [block for block in chart_slot.blocks if isinstance(block, HeadingBlock)]
    tables = [block for block in chart_slot.blocks if isinstance(block, TableBlock)]
    paragraphs = [block for block in chart_slot.blocks if isinstance(block, ParagraphBlock)]
    assert headings[0].inlines == (InlineText("Revenue trend"),)
    assert tables == [
        TableBlock(
            (
                ("Series", "Revenue", "Category", "Jan", "Value", "10"),
                ("Series", "Revenue", "Category", "Feb", "Value", "15"),
                ("Series", "Revenue", "Category", "Mar", "Value", "12"),
            ),
            header_rows=0,
        )
    ]
    paragraph_text = "\n".join(
        "".join(inline.text for inline in block.inlines if isinstance(inline, InlineText))
        for block in paragraphs
    )
    assert "Axis titles: Month; USD" in paragraph_text
    assert "Data labels: value" in paragraph_text
    assert "'Sales Data'!$A$2:$A$4" in paragraph_text
    assert "'Sales Data'!$B$2:$B$4" in paragraph_text


def test_chart_cache_wins_and_unsupported_references_are_preserved_without_access(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "chart-references.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quoted Sheet"
    for row in (("Label", "Value"), ("local", 1), ("ignored", 2)):
        sheet.append(row)
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D4")
    workbook.save(path)
    with ZipFile(path) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml")
    chart_xml = chart_xml.replace(
        b"<strRef><f>'Quoted Sheet'!B1</f></strRef>",
        (
            b"<strRef><f>'Quoted Sheet'!B1</f><strCache><ptCount val=\"1\"/>"
            b'<pt idx="0"><v>Cached series</v></pt></strCache></strRef>'
        ),
    )
    chart_xml = chart_xml.replace(
        b"<numRef><f>'Quoted Sheet'!$A$2:$A$3</f></numRef>",
        b"<numRef><f>[Book.xlsx]Quoted Sheet!$A$2:$A$3</f></numRef>",
    )
    chart_xml = chart_xml.replace(
        b"<numRef><f>'Quoted Sheet'!$B$2:$B$3</f></numRef>",
        b"<numRef><f>DynamicName</f></numRef>",
    )
    rewrite_xlsx(path, {"xl/charts/chart1.xml": chart_xml})

    def forbidden_network(*args: object, **kwargs: object) -> object:
        del args, kwargs
        raise AssertionError("chart references must not access the network")

    monkeypatch.setattr(socket, "create_connection", forbidden_network)
    monkeypatch.setattr(urllib.request, "urlopen", forbidden_network)
    document = extract_xlsx(
        path,
        preflight_xlsx(path),
        artifact_dir=tmp_path / "artifacts",
    )

    chart_slot = next(slot for slot in document.sheets[0].slots if isinstance(slot, XlsxChartSlot))
    assert (
        any(
            isinstance(block, HeadingBlock) and block.inlines == (InlineText("Cached series"),)
            for block in chart_slot.blocks
        )
        is False
    )
    rendered_facts = repr(chart_slot.blocks)
    assert "Cached series" in rendered_facts
    assert "[Book.xlsx]Quoted Sheet!$A$2:$A$3" in rendered_facts
    assert "DynamicName" in rendered_facts
    assert [warning.code for warning in document.warnings].count("xlsx_external_reference") == 2


def test_multiple_chart_reference_warnings_stay_bound_to_their_occurrence(
    tmp_path: Path,
) -> None:
    path = tmp_path / "chart-warning-occurrences.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(("Category", "Value"))
    sheet.append(("A", 1))
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=2))
    sheet.add_chart(chart, "D2")
    sheet.add_chart(deepcopy(chart), "D20")
    workbook.save(path)
    replacements: dict[str, bytes | None] = {}
    with ZipFile(path) as archive:
        for index, formula in ((1, b"DynamicOne"), (2, b"DynamicTwo")):
            part = f"xl/charts/chart{index}.xml"
            xml = archive.read(part)
            original = b"<numRef><f>'Data'!$B$2</f></numRef>"
            assert original in xml
            replacements[part] = xml.replace(
                original,
                b"<numRef><f>" + formula + b"</f></numRef>",
            )
    rewrite_xlsx(path, replacements)

    document = extract_xlsx(path, preflight_xlsx(path), artifact_dir=tmp_path / "artifacts")

    warnings = [
        warning for warning in document.warnings if warning.code == "xlsx_external_reference"
    ]
    assert [warning.message for warning in warnings] == [
        "Data!D2: chart reference was preserved without access: DynamicOne",
        "Data!D20: chart reference was preserved without access: DynamicTwo",
    ]


def test_semantic_chart_previews_and_requests_are_occurrence_independent(
    tmp_path: Path,
) -> None:
    path = tmp_path / "duplicate-charts.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Data"
    first.append(("X", "Y"))
    first.append((1, 3))
    first.append((2, 5))
    chart = ScatterChart()
    chart.title = "Relationship"
    cast(Any, chart.x_axis).title = "X"
    cast(Any, chart.y_axis).title = "Y"
    chart.series.append(
        Series(
            Reference(first, min_col=2, min_row=2, max_row=3),
            Reference(first, min_col=1, min_row=2, max_row=3),
            title="points",
        )
    )
    first.add_chart(chart, "D2")
    second = workbook.create_sheet("Other")
    second.add_chart(deepcopy(chart), "H8")
    workbook.save(path)
    with ZipFile(path) as archive:
        second_drawing = archive.read("xl/drawings/drawing2.xml")
    assert b'name="Chart 1"' in second_drawing
    rewrite_xlsx(
        path,
        {
            "xl/drawings/drawing2.xml": second_drawing.replace(
                b'name="Chart 1"',
                b'name="Different occurrence" descr="different alt"',
            )
        },
    )
    artifacts = tmp_path / "artifacts"

    document = extract_xlsx(path, preflight_xlsx(path), artifact_dir=artifacts)
    chart_slots = tuple(
        slot for sheet in document.sheets for slot in sheet.slots if isinstance(slot, XlsxChartSlot)
    )
    requests = build_xlsx_visual_requests(document, artifacts)

    assert [(slot.anchor, slot.content_sha256) for slot in chart_slots] == [
        ("D2", chart_slots[0].content_sha256),
        ("H8", chart_slots[0].content_sha256),
    ]
    assert chart_slots[0].artifact_name == chart_slots[1].artifact_name
    assert len(requests) == 1
    assert requests[0].digest == chart_slots[0].content_sha256
    assert requests[0].prompt == XLSX_CHART_VISION_PROMPT
    assert "趋势" in requests[0].prompt
    assert "关系" in requests[0].prompt
    assert "标注" in requests[0].prompt
    assert "含义" in requests[0].prompt
    assert "视觉解释" in requests[0].prompt


def test_without_artifact_directory_native_chart_text_survives_and_images_are_warned(
    tmp_path: Path,
) -> None:
    path = tmp_path / "native-only.xlsx"
    source_image = tmp_path / "source.png"
    _write_png(source_image)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Name", "Value"))
    sheet.append(("A", 1))
    chart = PieChart()
    chart.title = "Share"
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=2))
    sheet.add_chart(chart, "D2")
    sheet.add_image(Image(source_image), "J2")
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    assert not _visual_slots(document.sheets[0])
    native_text = repr(_native_slots(document.sheets[0]))
    assert "Share" in native_text
    assert "Series" in native_text
    assert "Image name:" in native_text
    assert [warning.code for warning in document.warnings].count(
        "xlsx_visual_artifact_unavailable"
    ) == 2


def test_image_visual_preparation_reuses_shared_safety_helper(tmp_path: Path) -> None:
    artifact_dir = tmp_path / "artifacts"
    output_dir = tmp_path / "prepared"
    artifact_dir.mkdir()
    output_dir.mkdir()
    disguised = artifact_dir / "xlsx-media.png"
    image = PILImage.new("RGB", (16, 16), "red")
    try:
        image.save(disguised, format="JPEG")
    finally:
        image.close()
    slot = XlsxImageSlot(
        source_index=1,
        anchor="A1",
        artifact_name=disguised.name,
        content_sha256="a" * 64,
    )

    with pytest.raises(DocumentTypeMismatchError, match="extension declares png"):
        prepare_xlsx_visual_artifact(slot, artifact_dir, output_dir, "prepared")


def test_supported_chart_families_emit_native_type_facts(tmp_path: Path) -> None:
    path = tmp_path / "chart-families.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(("Category", "Value", "X", "Y"))
    sheet.append(("A", 1, 1, 2))
    sheet.append(("B", 2, 2, 4))
    chart_types = (LineChart(), BarChart(), PieChart(), DoughnutChart())
    for index, chart in enumerate(chart_types, start=1):
        chart.title = type(chart).__name__
        chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
        sheet.add_chart(chart, f"F{index * 8}")
    scatter = ScatterChart()
    scatter.title = "ScatterChart"
    scatter.series.append(
        Series(
            Reference(sheet, min_col=4, min_row=2, max_row=3),
            Reference(sheet, min_col=3, min_row=2, max_row=3),
            title="points",
        )
    )
    sheet.add_chart(scatter, "F40")
    workbook.save(path)

    document = extract_xlsx(
        path,
        preflight_xlsx(path),
        artifact_dir=tmp_path / "artifacts",
    )

    chart_slots = [slot for slot in document.sheets[0].slots if isinstance(slot, XlsxChartSlot)]
    assert len(chart_slots) == 5
    assert [
        next(
            inline.text
            for block in slot.blocks
            if isinstance(block, ParagraphBlock)
            for inline in block.inlines
            if isinstance(inline, InlineText) and inline.text.startswith("Chart type:")
        )
        for slot in chart_slots
    ] == [
        "Chart type: line",
        "Chart type: bar",
        "Chart type: pie",
        "Chart type: doughnut",
        "Chart type: scatter",
    ]


def test_duplicate_embedded_media_share_raw_digest_and_one_request(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-images.xlsx"
    source_image = tmp_path / "shared.png"
    image_bytes = _write_png(source_image, color="green")
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.add_image(Image(source_image), "B2")
    second = workbook.create_sheet("Second")
    second.add_image(Image(source_image), "H9")
    workbook.save(path)
    artifacts = tmp_path / "artifacts"

    document = extract_xlsx(path, preflight_xlsx(path), artifact_dir=artifacts)
    image_slots = tuple(
        slot for sheet in document.sheets for slot in sheet.slots if isinstance(slot, XlsxImageSlot)
    )
    requests = build_xlsx_visual_requests(document, artifacts)

    assert [slot.anchor for slot in image_slots] == ["B2", "H9"]
    assert image_slots[0].content_sha256 == image_slots[1].content_sha256
    assert image_slots[0].artifact_name == image_slots[1].artifact_name
    assert (artifacts / image_slots[0].artifact_name).read_bytes() == image_bytes
    assert (
        len([request for request in requests if request.digest == image_slots[0].content_sha256])
        == 1
    )


def test_multilevel_category_cache_combines_levels_without_duplicate_index_failure(
    tmp_path: Path,
) -> None:
    path = tmp_path / "multilevel.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(("Category", "Value"))
    sheet.append(("Jan", 1))
    sheet.append(("Feb", 2))
    chart = LineChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
    sheet.add_chart(chart, "D2")
    workbook.save(path)
    with ZipFile(path) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml")
    original = b"<numRef><f>'Data'!$A$2:$A$3</f></numRef>"
    replacement = (
        b"<multiLvlStrRef><f>'Data'!$A$2:$A$3</f><multiLvlStrCache>"
        b'<ptCount val="2"/><lvl><pt idx="0"><v>Q1</v></pt>'
        b'<pt idx="1"><v>Q1</v></pt></lvl><lvl>'
        b'<pt idx="0"><v>Jan</v></pt><pt idx="1"><v>Feb</v></pt>'
        b"</lvl></multiLvlStrCache></multiLvlStrRef>"
    )
    assert original in chart_xml
    chart_xml = chart_xml.replace(original, replacement).replace(
        b"<chart><plotArea>",
        (b"<chart><title><tx><strRef><f>'Data'!$A$1</f></strRef></tx></title><plotArea>"),
    )
    rewrite_xlsx(path, {"xl/charts/chart1.xml": chart_xml})

    document = extract_xlsx(
        path,
        preflight_xlsx(path),
        artifact_dir=tmp_path / "artifacts",
    )

    slot = next(slot for slot in document.sheets[0].slots if isinstance(slot, XlsxChartSlot))
    heading = next(block for block in slot.blocks if isinstance(block, HeadingBlock))
    table = next(block for block in slot.blocks if isinstance(block, TableBlock))
    assert heading.inlines == (InlineText("Category"),)
    assert [row[3] for row in table.grid] == ["Q1 / Jan", "Q1 / Feb"]


def test_unsupported_chart_type_is_skipped_with_locatable_warning(tmp_path: Path) -> None:
    path = tmp_path / "unsupported-chart.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "kept"
    chart = LineChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    sheet.add_chart(chart, "D2")
    workbook.save(path)
    with ZipFile(path) as archive:
        chart_xml = archive.read("xl/charts/chart1.xml")
    assert b"<lineChart>" in chart_xml
    chart_xml = chart_xml.replace(b"<lineChart>", b"<areaChart>").replace(
        b"</lineChart>", b"</areaChart>"
    )
    rewrite_xlsx(path, {"xl/charts/chart1.xml": chart_xml})

    document = extract_xlsx(path, preflight_xlsx(path), artifact_dir=tmp_path / "artifacts")

    assert not any(isinstance(slot, XlsxChartSlot) for slot in document.sheets[0].slots)
    warning = next(
        warning for warning in document.warnings if warning.code == "xlsx_unsupported_object"
    )
    assert warning.message.startswith("Sheet!D2:")


@pytest.mark.parametrize(
    ("field", "limit", "message"),
    [
        ("drawing_objects", media_module.MAX_DRAWING_OBJECTS, "drawing object"),
        ("chart_cache_points", media_module.MAX_CHART_CACHE_POINTS, "chart cache point"),
    ],
)
def test_visual_outer_limits_fail_before_semantic_preview(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    field: str,
    limit: int,
    message: str,
) -> None:
    path = tmp_path / "bounded-visual.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "kept"
    workbook.save(path)
    preflight = preflight_xlsx(path)
    accepted = replace(
        preflight,
        usage=replace(preflight.usage, **{field: limit}),
    )
    bounded = replace(
        preflight,
        usage=replace(preflight.usage, **{field: limit + 1}),
    )

    def forbidden_preview(*args: object, **kwargs: object) -> bytes:
        del args, kwargs
        raise AssertionError("semantic preview must not run after an outer-limit failure")

    monkeypatch.setattr(media_module, "render_chart_semantic_preview", forbidden_preview)
    extract_xlsx(path, accepted, artifact_dir=tmp_path / "accepted-artifacts")
    with pytest.raises(LimitExceededError, match=message):
        extract_xlsx(path, bounded, artifact_dir=tmp_path / "artifacts")

    assert not (tmp_path / "artifacts").exists()


def test_extract_preserves_all_sheet_like_entries_states_and_empty_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "ordered.xlsx"
    _save_ordered_workbook(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    assert [
        (sheet.sheet_index, sheet.name, sheet.kind, sheet.state) for sheet in document.sheets
    ] == [
        (1, "Visible", XlsxSheetKind.WORKSHEET, XlsxSheetState.VISIBLE),
        (2, "Hidden", XlsxSheetKind.WORKSHEET, XlsxSheetState.HIDDEN),
        (3, "Very Hidden", XlsxSheetKind.WORKSHEET, XlsxSheetState.VERY_HIDDEN),
        (4, "Chart", XlsxSheetKind.CHARTSHEET, XlsxSheetState.VISIBLE),
    ]
    for expected_index, sheet in enumerate(document.sheets, start=1):
        prelude = _native_slots(sheet)[0]
        assert prelude.anchor == "A1"
        assert prelude.source_index == 0
        assert isinstance(prelude.blocks[0], MarkdownBlock)
        assert prelude.blocks[0].markdown == f"<!-- xlsx-sheet: {expected_index} -->"
        assert sheet.name not in prelude.blocks[0].markdown
        assert isinstance(prelude.blocks[1], HeadingBlock)
        assert isinstance(prelude.blocks[2], ParagraphBlock)
    assert len(_native_slots(document.sheets[1])) == 1
    assert len(_native_slots(document.sheets[2])) == 1
    assert len(_native_slots(document.sheets[3])) == 2


def test_extract_builds_tables_regions_merges_and_ignores_style_only_cells(
    tmp_path: Path,
) -> None:
    path = tmp_path / "regions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Regions"
    sheet.append(("Name", "Amount"))
    sheet.append(("A", 1))
    sheet.append(("B", 2))
    table = Table(displayName="Ledger", ref="A1:B3")
    table.headerRowCount = 1
    sheet.add_table(table)
    sheet["D1"] = "Merged"
    sheet.merge_cells("D1:E1")
    sheet["D2"] = "left"
    sheet["E2"] = "right"
    sheet["H1"] = "first"
    sheet["H3"] = "second"
    sheet["J1"].font = Font(bold=True, color="FF0000")
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    slots = _native_slots(document.sheets[0])
    assert [slot.anchor for slot in slots] == ["A1", "A1:B3", "D1:E2", "H1", "H3"]
    assert [slot.source_index for slot in slots] == list(range(5))
    assert isinstance(slots[1].blocks[1], TableBlock)
    assert slots[1].blocks[1].header_rows == 1
    assert slots[1].blocks[1].grid == (("Name", "Amount"), ("A", "1"), ("B", "2"))
    assert isinstance(slots[2].blocks[1], SpannedTableBlock)
    assert slots[2].blocks[1].header_rows == 0
    assert [(cell.row_span, cell.column_span, cell.text) for cell in slots[2].blocks[1].cells] == [
        (1, 2, "Merged"),
        (1, 1, "left"),
        (1, 1, "right"),
    ]
    assert isinstance(slots[3].blocks[1], TableBlock)
    assert slots[3].blocks[1].header_rows == 0
    assert all(
        "Regions" not in block.markdown
        for slot in slots
        for block in slot.blocks
        if isinstance(block, MarkdownBlock)
    )
    assert all(slot.anchor != "J1" for slot in slots)


def test_excel_table_header_row_count_zero_is_not_promoted_to_header(tmp_path: Path) -> None:
    path = tmp_path / "headerless.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "one"
    sheet["A2"] = "two"
    table = Table(displayName="Headerless", ref="A1:A2")
    table.headerRowCount = 0
    sheet.add_table(table)
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    table_block = _native_slots(document.sheets[0])[1].blocks[1]
    assert isinstance(table_block, TableBlock)
    assert table_block.header_rows == 0


def test_extracted_native_blocks_render_to_stable_markdown(tmp_path: Path) -> None:
    path = tmp_path / "golden.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ledger"
    sheet.append(("Item", "Amount"))
    sheet.append(("Book", 12.5))
    sheet["B2"].number_format = "$0.00"
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))
    blocks = tuple(
        block
        for extracted_sheet in document.sheets
        for slot in extracted_sheet.slots
        if isinstance(slot, XlsxNativeSlot)
        for block in slot.blocks
    )
    rendered = render_markdown(
        ParsedDocument(DocumentType.XLSX, blocks, document.warnings),
        max_output_chars=10_000,
    )

    assert rendered.markdown == (
        "<!-- xlsx-sheet: 1 -->\n\n"
        "# Ledger\n\n"
        "Sheet state: visible\n\n"
        "<!-- xlsx-region: sheet=1 range=A1:B2 object=1 -->\n\n"
        "<table>\n"
        "<tbody>\n"
        "<tr><td>Item</td><td>Amount</td></tr>\n"
        "<tr><td>Book</td><td>$12.50</td></tr>\n"
        "</tbody>\n"
        "</table>\n"
    )


def test_extract_rechecks_component_bounding_box_before_materializing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "bounded.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "a"
    sheet["B1"] = "b"
    sheet["B2"] = "c"
    workbook.save(path)
    index = preflight_xlsx(path)
    monkeypatch.setattr(extract_module, "MAX_MATERIALIZED_GRID_CELLS", 3)

    with pytest.raises(LimitExceededError, match="materialized grid"):
        extract_xlsx(path, index)


def test_extract_rechecks_materialized_grid_across_all_sheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "bounded-workbook.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.append(("a", "b"))
    second = workbook.create_sheet("Second")
    second.append(("c", "d"))
    workbook.save(path)
    index = preflight_xlsx(path)
    monkeypatch.setattr(extract_module, "MAX_MATERIALIZED_GRID_CELLS", 3)

    with pytest.raises(LimitExceededError, match="materialized grid"):
        extract_xlsx(path, index)


def test_extract_aggregates_number_format_warnings_after_twenty_coordinates(
    tmp_path: Path,
) -> None:
    path = tmp_path / "warnings.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Warnings"
    for row in range(1, 23):
        cell = sheet.cell(row, 1, row)
        cell.number_format = "0.00E+00"
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    warnings = [
        warning for warning in document.warnings if warning.code == "xlsx_unsupported_number_format"
    ]
    assert len(warnings) == 21
    assert warnings[0].message.startswith("Warnings!A1:")
    assert warnings[19].message.startswith("Warnings!A20:")
    assert warnings[20].message == "2 additional xlsx_unsupported_number_format warnings suppressed"


def test_extract_falls_back_when_conditional_rule_can_change_number_format(
    tmp_path: Path,
) -> None:
    path = tmp_path / "conditional-number-format.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 1234.5
    sheet.conditional_formatting.add(
        "A1",
        Rule(
            type="expression",
            formula=("1",),
            dxf=DifferentialStyle(numFmt=NumberFormat(numFmtId=164, formatCode="$0.00")),
        ),
    )
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    region = _native_slots(document.sheets[0])[1]
    assert isinstance(region.blocks[1], TableBlock)
    assert region.blocks[1].grid == (("1234.5",),)
    assert [warning.code for warning in document.warnings] == ["xlsx_unsupported_number_format"]


def test_extract_loads_full_mode_openpyxl_once_with_links_disabled(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "once.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "value"
    workbook.save(path)
    index = preflight_xlsx(path)
    calls: list[tuple[Path, dict[str, object]]] = []
    original = extract_module.openpyxl.load_workbook

    def recording_loader(filename: Path, **kwargs: object) -> object:
        calls.append((filename, kwargs))
        return original(filename, **kwargs)

    monkeypatch.setattr(extract_module.openpyxl, "load_workbook", recording_loader)

    extract_xlsx(path, index)

    assert calls == [
        (
            path,
            {
                "read_only": False,
                "data_only": False,
                "rich_text": False,
                "keep_links": False,
            },
        )
    ]


def test_formula_sidecar_prefers_cache_and_distinguishes_missing_empty_and_special_formulas(
    tmp_path: Path,
) -> None:
    path = tmp_path / "formulas.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formulas"
    for row in range(1, 9):
        sheet.cell(row, 1, row)
        sheet.cell(row, 2, f"=A{row}*2")
    sheet["B1"].number_format = "$#,##0.00"
    workbook.save(path)
    with ZipFile(path) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml")
    replacements = {
        b'<c r="B1" s="1"><f>A1*2</f><v></v></c>': b'<c r="B1" s="1"><f>A1*2</f><v>2</v></c>',
        b'<c r="B2"><f>A2*2</f><v></v></c>': b'<c r="B2"><f>A2*2</f></c>',
        b'<c r="B3"><f>A3*2</f><v></v></c>': b'<c r="B3"><f>A3*2</f><v></v></c>',
        b'<c r="B4"><f>A4*2</f><v></v></c>': b'<c r="B4"><f>[1]Sheet1!A1</f></c>',
        b'<c r="B5"><f>A5*2</f><v></v></c>': (
            b'<c r="B5"><f t="array" ref="B5:C5">SUM(A5:A6)</f><v>10</v></c>'
        ),
        b'<c r="B6"><f>A6*2</f><v></v></c>': (
            b'<c r="B6"><f t="dataTable" ref="B6:C7" dt2D="1" r1="A1" r2="A2"></f><v></v></c>'
        ),
        b'<c r="B7"><f>A7*2</f><v></v></c>': (
            b'<c r="B7"><f t="shared" ref="B7:B8" si="0">A7*2</f></c>'
        ),
        b'<c r="B8"><f>A8*2</f><v></v></c>': b'<c r="B8"><f t="shared" si="0"></f></c>',
    }
    for old, new in replacements.items():
        assert old in xml
        xml = xml.replace(old, new)
    rewrite_xlsx(path, {"xl/worksheets/sheet1.xml": xml})

    document = extract_xlsx(path, preflight_xlsx(path))

    text_by_anchor = {
        slot.anchor: "\n".join(cell for row in slot.blocks[1].grid for cell in row)
        for slot in _native_slots(document.sheets[0])[1:]
        if isinstance(slot.blocks[1], TableBlock)
    }
    all_text = "\n".join(text_by_anchor.values())
    assert "$2.00" in all_text
    assert "=A2*2" in all_text
    assert "=A3*2" not in all_text
    assert "=[1]Sheet1!A1" in all_text
    assert "Array/spill formula B5:C5: =SUM(A5:A6); saved value: 10" in all_text
    assert "Data-table formula B6:C7 (dt2D=1, r1=A1, r2=A2)" in all_text
    assert "=A7*2" in all_text
    assert "=A8*2" in all_text
    codes = [warning.code for warning in document.warnings]
    assert codes.count("xlsx_formula_cache_missing") == 4
    assert codes.count("xlsx_data_table_formula") == 1
    assert codes.count("xlsx_external_reference") == 1


def test_repeated_extraction_is_deterministic(tmp_path: Path) -> None:
    path = tmp_path / "repeat.xlsx"
    _save_ordered_workbook(path)
    index = preflight_xlsx(path)

    assert extract_xlsx(path, index) == extract_xlsx(path, index)


def test_extracts_comments_threaded_text_boxes_links_and_headers_in_stable_order(
    tmp_path: Path,
) -> None:
    path = tmp_path / "text-objects.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Objects"
    sheet.append(("HTTP [docs]", "HTTPS", "Mail", "Jump"))
    workbook.save(path)
    with ZipFile(path) as archive:
        worksheet_xml = _with_relationship_namespace(archive.read("xl/worksheets/sheet1.xml"))
        workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
    worksheet_xml = _append_before(
        worksheet_xml,
        b"</worksheet>",
        (
            '<hyperlinks><hyperlink ref="A1" r:id="rIdHttp" display="HTTP [docs]"/>'
            '<hyperlink ref="B1" r:id="rIdHttps" display="HTTPS"/>'
            '<hyperlink ref="C1" r:id="rIdMail" display="Mail"/>'
            '<hyperlink ref="D1" location="Objects!A1" display="Jump"/></hyperlinks>'
            '<drawing r:id="rIdDrawing"/>'
            "<headerFooter>"
            "<oddHeader>&amp;LOdd header left &amp;&amp; kept &amp;P &amp;N"
            "&amp;C&amp;D &amp;T&amp;R&amp;&quot;Arial&quot;&amp;KFF0000&amp;12&amp;B"
            "Odd header right &amp;F &amp;Z &amp;A &amp;G</oddHeader>"
            "<oddFooter>&amp;LOdd footer left&amp;COdd footer center&amp;ROdd footer right"
            "</oddFooter>"
            "<evenHeader>&amp;LEven header left&amp;CEven header center&amp;REven header right"
            "</evenHeader>"
            "<evenFooter>&amp;LEven footer left&amp;CEven footer center&amp;REven footer right"
            "</evenFooter>"
            "<firstHeader>&amp;LFirst header left&amp;CFirst header center&amp;RFirst header right"
            "</firstHeader>"
            "<firstFooter>&amp;LFirst footer left&amp;CFirst footer center&amp;RFirst footer right"
            "</firstFooter>"
            "</headerFooter>"
        ),
    )
    workbook_rels = _append_before(
        workbook_rels,
        b"</Relationships>",
        (
            f'<Relationship Id="rIdPersons" Type="{THREADED_REL_NS}/person" '
            'Target="persons/person.xml"/>'
        ),
    )
    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": worksheet_xml,
            "xl/worksheets/_rels/sheet1.xml.rels": _relationships(
                ("rIdHttp", f"{OFFICE_REL_NS}/hyperlink", "http://example.test/a(b)", True),
                ("rIdHttps", f"{OFFICE_REL_NS}/hyperlink", "https://example.test/two", True),
                ("rIdMail", f"{OFFICE_REL_NS}/hyperlink", "mailto:team@example.test", True),
                ("rIdComment", f"{OFFICE_REL_NS}/comments", "../comments1.xml", False),
                (
                    "rIdThreaded",
                    f"{THREADED_REL_NS}/threadedComment",
                    "../threadedComments/threadedComment1.xml",
                    False,
                ),
                ("rIdDrawing", f"{OFFICE_REL_NS}/drawing", "../drawings/drawing1.xml", False),
            ),
            "xl/comments1.xml": (
                f'<comments xmlns="{SHEET_NS}"><authors><author>Alice</author></authors>'
                '<commentList><comment ref="B2" authorId="0"><text><t>Classic note</t>'
                "</text></comment></commentList></comments>"
            ).encode(),
            "xl/threadedComments/threadedComment1.xml": (
                f'<ThreadedComments xmlns="{THREADED_NS}"><threadedComment ref="C3" '
                'personId="person-1"><text>Threaded reply</text></threadedComment>'
                "</ThreadedComments>"
            ).encode(),
            "xl/persons/person.xml": (
                f'<personList xmlns="{THREADED_NS}"><person id="person-1" '
                'displayName="Bob" userId="bob@example.test" providerId="None"/>'
                "</personList>"
            ).encode(),
            "xl/_rels/workbook.xml.rels": workbook_rels,
            "xl/drawings/drawing1.xml": (
                f'<xdr:wsDr xmlns:xdr="{DRAWING_NS}" xmlns:a="{DRAWING_MAIN_NS}">'
                "<xdr:oneCellAnchor><xdr:from><xdr:col>4</xdr:col><xdr:row>1</xdr:row>"
                '</xdr:from><xdr:ext cx="1" cy="1"/><xdr:sp><xdr:nvSpPr>'
                '<xdr:cNvPr id="2" name="Callout 1" descr="Shape details" '
                'title="Shape title"/><xdr:cNvSpPr/></xdr:nvSpPr><xdr:spPr/>'
                "<xdr:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Box &amp; text</a:t>"
                "</a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData/>"
                "</xdr:oneCellAnchor></xdr:wsDr>"
            ).encode(),
        },
    )

    document = extract_xlsx(path, preflight_xlsx(path))

    slots = _native_slots(document.sheets[0])
    assert [slot.source_index for slot in slots] == list(range(len(slots)))
    assert slots[0].blocks[0] == MarkdownBlock("<!-- xlsx-sheet: 1 -->")
    links = [
        inline
        for slot in slots
        for block in slot.blocks
        if isinstance(block, ParagraphBlock)
        for inline in block.inlines
        if isinstance(inline, InlineLink)
    ]
    assert [(link.label, link.target) for link in links] == [
        ("HTTP [docs]", "http://example.test/a(b)"),
        ("HTTPS", "https://example.test/two"),
        ("Mail", "mailto:team@example.test"),
        ("Jump", "#Objects!A1"),
    ]
    all_text = [_paragraph_text(slot) for slot in slots]
    assert "Comment by Alice: Classic note" in all_text
    assert "Threaded comment by Bob: Threaded reply" in all_text
    assert any(
        text == "Text box: Box & text\nShape name: Callout 1\nAlt text: Shape details\n"
        "Alt title: Shape title"
        for text in all_text
    )
    header_text = [text for text in all_text if text.startswith(("Odd ", "Even ", "First "))]
    assert header_text == [
        "Odd header left: Odd header left & kept {page} {pages}",
        "Odd header center: {date} {time}",
        "Odd header right: Odd header right {file} {path} {sheet}",
        "Odd footer left: Odd footer left",
        "Odd footer center: Odd footer center",
        "Odd footer right: Odd footer right",
        "Even header left: Even header left",
        "Even header center: Even header center",
        "Even header right: Even header right",
        "Even footer left: Even footer left",
        "Even footer center: Even footer center",
        "Even footer right: Even footer right",
        "First header left: First header left",
        "First header center: First header center",
        "First header right: First header right",
        "First footer left: First footer left",
        "First footer center: First footer center",
        "First footer right: First footer right",
    ]
    assert all("Arial" not in text and "FF0000" not in text for text in header_text)
    assert [warning.code for warning in document.warnings] == ["xlsx_unsupported_object"]
    assert "sheet=1 anchor=A1" in document.warnings[0].message
    assert "header/footer image field" in document.warnings[0].message
    assert all(
        "Objects" not in block.markdown
        for slot in slots[1:]
        for block in slot.blocks
        if isinstance(block, MarkdownBlock)
    )
    rendered = render_markdown(
        ParsedDocument(
            DocumentType.XLSX,
            tuple(block for slot in slots for block in slot.blocks),
            document.warnings,
        ),
        max_output_chars=100_000,
    )
    assert "[HTTP \\[docs\\]](http://example.test/a\\(b\\))" in rendered.markdown


def test_standard_note_vml_presentation_is_not_reported_as_lost_text(tmp_path: Path) -> None:
    path = tmp_path / "standard-note.xlsx"
    workbook = Workbook()
    workbook.active["B2"].comment = Comment("A standard note", "Reviewer")
    workbook.save(path)

    document = extract_xlsx(path, preflight_xlsx(path))

    paragraphs = [_paragraph_text(slot) for slot in _native_slots(document.sheets[0])]
    assert "Comment by Reviewer: A standard note" in paragraphs
    assert not any(
        warning.code == "xlsx_unsupported_object" and "vmlDrawing" in warning.message
        for warning in document.warnings
    )


def test_unsafe_hyperlinks_and_remote_data_are_plain_text_and_never_accessed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "external.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("Script", "File", "Book"))
    workbook.save(path)
    with ZipFile(path) as archive:
        worksheet_xml = _with_relationship_namespace(archive.read("xl/worksheets/sheet1.xml"))
        workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
    worksheet_xml = _append_before(
        worksheet_xml,
        b"</worksheet>",
        (
            '<hyperlinks><hyperlink ref="A1" r:id="rIdScript"/>'
            '<hyperlink ref="B1" r:id="rIdFile" display="File"/>'
            '<hyperlink ref="C1" r:id="rIdBook" display="Book"/></hyperlinks>'
        ),
    )
    workbook_rels = _append_before(
        workbook_rels,
        b"</Relationships>",
        (
            f'<Relationship Id="rIdExternal" Type="{OFFICE_REL_NS}/externalLink" '
            'Target="externalLinks/externalLink1.xml"/>'
            f'<Relationship Id="rIdConnections" Type="{OFFICE_REL_NS}/connections" '
            'Target="connections.xml"/>'
        ),
    )
    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": worksheet_xml,
            "xl/worksheets/_rels/sheet1.xml.rels": _relationships(
                ("rIdScript", f"{OFFICE_REL_NS}/hyperlink", "javascript:alert(1)", True),
                ("rIdFile", f"{OFFICE_REL_NS}/hyperlink", "file:///tmp/private.xlsx", True),
                ("rIdBook", f"{OFFICE_REL_NS}/hyperlink", "../other.xlsx", True),
            ),
            "xl/_rels/workbook.xml.rels": workbook_rels,
            "xl/externalLinks/externalLink1.xml": (f'<externalLink xmlns="{SHEET_NS}"/>').encode(),
            "xl/externalLinks/_rels/externalLink1.xml.rels": _relationships(
                (
                    "rIdPath",
                    f"{OFFICE_REL_NS}/externalLinkPath",
                    "https://remote.example.test/book.xlsx",
                    True,
                ),
            ),
            "xl/connections.xml": (
                f'<connections xmlns="{SHEET_NS}"><connection id="1" name="Remote" '
                'type="5" sourceFile="https://remote.example.test/data.csv"/>'
                "</connections>"
            ).encode(),
        },
    )

    def deny_network(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"network access attempted: {args!r} {kwargs!r}")

    monkeypatch.setattr(socket, "create_connection", deny_network)
    monkeypatch.setattr(urllib.request, "urlopen", deny_network)
    for module_name in ("httpx", "requests"):
        module = pytest.importorskip(module_name)
        monkeypatch.setattr(module, "get", deny_network)

    document = extract_xlsx(path, preflight_xlsx(path))

    paragraphs = [
        block
        for slot in _native_slots(document.sheets[0])
        for block in slot.blocks
        if isinstance(block, ParagraphBlock)
    ]
    assert not any(
        isinstance(inline, InlineLink) for paragraph in paragraphs for inline in paragraph.inlines
    )
    plain = "\n".join(
        inline.text
        for paragraph in paragraphs
        for inline in paragraph.inlines
        if isinstance(inline, InlineText)
    )
    assert "Script (javascript:alert(1))" in plain
    assert "File (file:///tmp/private.xlsx)" in plain
    assert "Book (../other.xlsx)" in plain
    assert "https://remote.example.test/book.xlsx" in plain
    assert "https://remote.example.test/data.csv" in plain
    external_warnings = [
        warning for warning in document.warnings if warning.code == "xlsx_external_reference"
    ]
    assert len(external_warnings) == 5
    assert all("sheet=1" in warning.message for warning in external_warnings)


def test_unsupported_xlsx_objects_are_locatable_and_aggregated(tmp_path: Path) -> None:
    path = tmp_path / "unsupported.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "kept"
    workbook.save(path)
    with ZipFile(path) as archive:
        worksheet_xml = _with_relationship_namespace(archive.read("xl/worksheets/sheet1.xml"))
    worksheet_xml = _append_before(
        worksheet_xml,
        b"</worksheet>",
        '<extLst><ext uri="vendor-extension"><vendor xmlns="urn:vendor"/></ext></extLst>',
    )
    unsupported = (
        ("rIdSmartArt", f"{OFFICE_REL_NS}/diagramData", "../diagrams/data1.xml", False),
        ("rIdOle", f"{OFFICE_REL_NS}/oleObject", "../embeddings/ole1.bin", False),
        ("rIdActiveX", f"{OFFICE_REL_NS}/activeXControl", "../activeX/activeX1.xml", False),
        ("rIdControl", f"{OFFICE_REL_NS}/control", "../controls/control1.xml", False),
        ("rIdVml", f"{OFFICE_REL_NS}/vmlDrawing", "../drawings/vmlDrawing1.vml", False),
    )
    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": worksheet_xml,
            "xl/worksheets/_rels/sheet1.xml.rels": _relationships(*unsupported),
            "xl/diagrams/data1.xml": b'<dgm:dataModel xmlns:dgm="urn:diagram"/>',
            "xl/embeddings/ole1.bin": b"ole",
            "xl/activeX/activeX1.xml": b'<ax:ocx xmlns:ax="urn:activex"/>',
            "xl/controls/control1.xml": b'<control xmlns="urn:control"/>',
            "xl/drawings/vmlDrawing1.vml": (
                b'<xml xmlns:v="urn:schemas-microsoft-com:vml"><v:shape>'
                b"<v:textbox>VML text</v:textbox></v:shape></xml>"
            ),
        },
    )

    document = extract_xlsx(path, preflight_xlsx(path))

    warnings = [
        warning for warning in document.warnings if warning.code == "xlsx_unsupported_object"
    ]
    assert len(warnings) == 6
    assert all("sheet=1" in warning.message for warning in warnings)
    assert all("object=" in warning.message for warning in warnings)
    assert any("diagramData" in warning.message for warning in warnings)
    assert any("oleObject" in warning.message for warning in warnings)
    assert any("activeXControl" in warning.message for warning in warnings)
    assert any("control" in warning.message for warning in warnings)
    assert any("vmlDrawing" in warning.message for warning in warnings)
    assert any("vendor extension" in warning.message for warning in warnings)


def test_unsupported_object_warnings_keep_twenty_then_summarize(tmp_path: Path) -> None:
    path = tmp_path / "many-unsupported.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "kept"
    workbook.save(path)
    with ZipFile(path) as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
    extensions = "".join(
        f'<ext uri="vendor-{index}"><vendor xmlns="urn:vendor"/></ext>' for index in range(22)
    )
    worksheet_xml = _append_before(
        worksheet_xml,
        b"</worksheet>",
        f"<extLst>{extensions}</extLst>",
    )
    rewrite_xlsx(path, {"xl/worksheets/sheet1.xml": worksheet_xml})

    document = extract_xlsx(path, preflight_xlsx(path))

    warnings = [
        warning for warning in document.warnings if warning.code == "xlsx_unsupported_object"
    ]
    assert len(warnings) == 21
    assert warnings[-1].message == "2 additional xlsx_unsupported_object warnings suppressed"


@pytest.mark.parametrize("relationship_id", ["missing", "wrong-type"])
def test_hyperlink_relationship_must_exist_with_the_expected_type(
    tmp_path: Path,
    relationship_id: str,
) -> None:
    path = tmp_path / f"bad-link-{relationship_id}.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "link"
    workbook.save(path)
    with ZipFile(path) as archive:
        worksheet_xml = _with_relationship_namespace(archive.read("xl/worksheets/sheet1.xml"))
    worksheet_xml = _append_before(
        worksheet_xml,
        b"</worksheet>",
        f'<hyperlinks><hyperlink ref="A1" r:id="{relationship_id}"/></hyperlinks>',
    )
    rewrite_xlsx(
        path,
        {
            "xl/worksheets/sheet1.xml": worksheet_xml,
            "xl/worksheets/_rels/sheet1.xml.rels": _relationships(
                (
                    "wrong-type",
                    f"{OFFICE_REL_NS}/comments",
                    "../comments1.xml",
                    False,
                ),
            ),
            "xl/comments1.xml": (
                f'<comments xmlns="{SHEET_NS}"><authors/><commentList/></comments>'
            ).encode(),
        },
    )

    with pytest.raises(CorruptDocumentError, match="relationship"):
        preflight_xlsx(path)
