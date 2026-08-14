from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from typing import Literal
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
OFFICE_DOCUMENT_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
)


def minimal_xlsx_entries(
    *,
    include_content_types: bool = True,
    include_root_relationships: bool = True,
    include_workbook: bool = True,
    workbook_content_type: str = XLSX_CONTENT_TYPE,
    root_target: str = "xl/workbook.xml",
    root_target_mode: str | None = None,
) -> list[tuple[str, bytes]]:
    entries: list[tuple[str, bytes]] = []
    if include_content_types:
        entries.append(
            (
                "[Content_Types].xml",
                f"""<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Override PartName="/xl/workbook.xml" ContentType="{workbook_content_type}"/>
</Types>
""".encode(),
            )
        )
    if include_root_relationships:
        target_mode = f' TargetMode="{root_target_mode}"' if root_target_mode else ""
        entries.append(
            (
                "_rels/.rels",
                f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="{OFFICE_DOCUMENT_RELATIONSHIP}"
   Target="{root_target}"{target_mode}/>
</Relationships>
""".encode(),
            )
        )
    if include_workbook:
        entries.append(
            (
                "xl/workbook.xml",
                b"<workbook xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'/>",
            )
        )
    return entries


def xlsx_bytes(
    *,
    include_content_types: bool = True,
    include_root_relationships: bool = True,
    include_workbook: bool = True,
    workbook_content_type: str = XLSX_CONTENT_TYPE,
    root_target: str = "xl/workbook.xml",
    root_target_mode: str | None = None,
) -> bytes:
    output = io.BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for name, data in minimal_xlsx_entries(
            include_content_types=include_content_types,
            include_root_relationships=include_root_relationships,
            include_workbook=include_workbook,
            workbook_content_type=workbook_content_type,
            root_target=root_target,
            root_target_mode=root_target_mode,
        ):
            archive.writestr(name, data)
    return output.getvalue()


def write_xlsx(
    path: Path,
    *,
    include_content_types: bool = True,
    include_root_relationships: bool = True,
    include_workbook: bool = True,
    workbook_content_type: str = XLSX_CONTENT_TYPE,
    root_target: str = "xl/workbook.xml",
    root_target_mode: str | None = None,
) -> None:
    path.write_bytes(
        xlsx_bytes(
            include_content_types=include_content_types,
            include_root_relationships=include_root_relationships,
            include_workbook=include_workbook,
            workbook_content_type=workbook_content_type,
            root_target=root_target,
            root_target_mode=root_target_mode,
        )
    )


def write_structured_xlsx(
    path: Path,
    *,
    sheets: tuple[
        tuple[
            str,
            Literal["worksheet", "chartsheet"],
            Literal["visible", "hidden", "veryHidden"],
            str | None,
            tuple[str, ...],
        ],
        ...,
    ],
) -> None:
    workbook_sheets: list[str] = []
    relationships: list[str] = []
    entries = minimal_xlsx_entries(include_workbook=False)
    for index, (name, kind, state, dimension, cells) in enumerate(sheets, start=1):
        relationship_id = f"rId{index}"
        workbook_sheets.append(
            f'<sheet name="{name}" sheetId="{index}" state="{state}" r:id="{relationship_id}"/>'
        )
        folder = "worksheets" if kind == "worksheet" else "chartsheets"
        relationship_type = (
            f"http://schemas.openxmlformats.org/officeDocument/2006/relationships/{kind}"
        )
        relationships.append(
            f'<Relationship Id="{relationship_id}" Type="{relationship_type}" '
            f'Target="{folder}/sheet{index}.xml"/>'
        )
        if kind == "worksheet":
            dimension_xml = f'<dimension ref="{dimension}"/>' if dimension else ""
            cell_xml = "".join(f'<c r="{cell}"><v>1</v></c>' for cell in cells)
            entries.append(
                (
                    f"xl/{folder}/sheet{index}.xml",
                    (
                        '<?xml version="1.0" encoding="UTF-8"?>'
                        '<worksheet xmlns="http://schemas.openxmlformats.org/'
                        'spreadsheetml/2006/main">'
                        f"{dimension_xml}<sheetData><row>{cell_xml}</row></sheetData>"
                        "</worksheet>"
                    ).encode(),
                )
            )
        else:
            entries.append(
                (
                    f"xl/{folder}/sheet{index}.xml",
                    b'<chartsheet xmlns="http://schemas.openxmlformats.org/'
                    b'spreadsheetml/2006/main"/>',
                )
            )
    entries.extend(
        [
            (
                "xl/workbook.xml",
                (
                    '<?xml version="1.0" encoding="UTF-8"?>'
                    '<workbook xmlns="http://schemas.openxmlformats.org/'
                    'spreadsheetml/2006/main" '
                    'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/'
                    'relationships"><sheets>'
                    f"{''.join(workbook_sheets)}"
                    "</sheets></workbook>"
                ).encode(),
            ),
            (
                "xl/_rels/workbook.xml.rels",
                (
                    '<?xml version="1.0" encoding="UTF-8"?>'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
                    f'relationships">{"".join(relationships)}</Relationships>'
                ).encode(),
            ),
        ]
    )
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, data in entries:
            archive.writestr(name, data)


def rewrite_xlsx(
    path: Path,
    replacements: dict[str, bytes | None],
) -> None:
    with ZipFile(path) as archive:
        entries = {info.filename: archive.read(info) for info in archive.infolist()}
    for name, data in replacements.items():
        if data is None:
            entries.pop(name, None)
        else:
            entries[name] = data
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)


def write_public_contract_xlsx(path: Path) -> None:
    workbook = Workbook()
    ledger = workbook.active
    ledger.title = "Ledger"
    ledger.append(("Item", "Amount", "Date", "Formula", "Scientific"))
    ledger.append(("Book", 1234.5, date(2026, 8, 14), "=B2*2", 1200))
    ledger["B2"].number_format = "$#,##0.00"
    ledger["C2"].number_format = "yyyy-mm-dd"
    ledger["E2"].number_format = "0.00E+00"
    ledger["A4"] = "Merged note"
    ledger.merge_cells("A4:B4")

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "hidden value"
    very_hidden = workbook.create_sheet("Very Hidden")
    very_hidden.sheet_state = "veryHidden"
    very_hidden["A1"] = "very hidden value"
    workbook.create_sheet("Empty")
    workbook.save(path)
    workbook.close()

    with ZipFile(path) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")
    formula = b'<c r="D2"><f>B2*2</f><v></v></c>'
    assert formula in worksheet
    rewrite_xlsx(
        path,
        {"xl/worksheets/sheet1.xml": worksheet.replace(formula, formula.replace(b"<v></v>", b""))},
    )
